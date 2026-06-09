"""

SIMULADOR CLÍNICA GINECOLÓGICA — VERSIÓN PARALELIZADA

Cambios respecto a la línea base:
  - main() usa ProcessPoolExecutor para correr réplicas en paralelo.
  - _run_worker() es un wrapper de nivel módulo (pickleable) que reconstruye
    SimConfig desde un dict antes de llamar run_once(), evitando problemas de
    pickle con dataclasses que tienen field(default_factory=...).
  - El número de workers se controla con CFG.n_workers (0 = auto = nCPUs).
  - El plot y los prints siguen corriendo en el proceso principal.
  - Todo lo demás es idéntico a la línea base.

"""

import matplotlib
try:
    matplotlib.use("TkAgg")
except Exception:
    matplotlib.use("Agg")

import simpy
import random
import numpy as np
import logging
import abc
from collections import deque, defaultdict
from dataclasses import dataclass, field
from enum import Enum
from typing import FrozenSet, Optional
import itertools
import statistics as stats
import matplotlib.pyplot as plt

# ── Paralelismo ──────────────────────────────────────────────────────────────
import concurrent.futures
import multiprocessing
import dataclasses

# ========================
# Logging estructurado
# ========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
log = logging.getLogger("clinica")

# ========================
# Priority enum
# ========================
class Priority(str, Enum):
    HIGH = "high"
    MID  = "mid"
    LOW  = "low"

# ========================
# Clase Calendar
# ========================
@dataclass(frozen=True)
class Calendar:
    start_hour: int
    end_hour:   int
    weekdays:   FrozenSet[int] = frozenset(range(5))

    def is_work_minute(self, t: float) -> bool:
        day = int(t // (24 * 60))
        dow = day % 7
        if dow not in self.weekdays:
            return False
        m = t % (24 * 60)
        return (self.start_hour * 60) <= m < (self.end_hour * 60)

    def next_work_minute(self, t: float) -> float:
        if self.is_work_minute(t):
            return t
        day = int(t // (24 * 60))
        while True:
            dow = day % 7
            if dow in self.weekdays:
                start = day * 24 * 60 + self.start_hour * 60
                if t <= start:
                    return float(start)
            day += 1
            t = day * 24 * 60

    def end_of_day(self, t: float) -> float:
        day = int(t // (24 * 60))
        return float(day * 24 * 60 + self.end_hour * 60)

    def total_work_minutes_until(self, horizon_min: float) -> float:
        total = 0.0
        day_idx = 0
        while True:
            day_start = day_idx * 24 * 60
            if day_start >= horizon_min:
                break
            dow = day_idx % 7
            if dow in self.weekdays:
                start = day_start + self.start_hour * 60
                end   = day_start + self.end_hour   * 60
                if start < horizon_min:
                    total += max(0.0, min(end, horizon_min) - start)
            day_idx += 1
        return total

    def consume(self, env, minutes: float):
        remaining = float(minutes)
        while remaining > 1e-9:
            t1 = self.next_work_minute(env.now)
            if t1 > env.now:
                yield env.timeout(t1 - env.now)
            end_today = self.end_of_day(env.now)
            dt = min(remaining, end_today - env.now)
            if dt <= 0:
                yield env.timeout(self.next_work_minute(env.now + 1) - env.now)
                continue
            yield env.timeout(dt)
            remaining -= dt

    def add_workdays(self, t0: float, workdays: int) -> float:
        t = t0
        if not self.is_work_minute(t):
            t = self.next_work_minute(t)
        days = 0
        while days < workdays:
            end_today = self.end_of_day(t)
            t = self.next_work_minute(end_today + 1)
            days += 1
        return t

    def workdays_before(self, t0: float, workdays: int) -> float:
        t = t0
        if not self.is_work_minute(t):
            t = self.next_work_minute(t)
        days = 0
        while days < workdays:
            day = int(t // (24 * 60))
            start_today = day * 24 * 60 + self.start_hour * 60
            if t > start_today:
                t = start_today
            else:
                day -= 1
                while (day % 7) not in self.weekdays:
                    day -= 1
                t = day * 24 * 60 + self.end_hour * 60
                days += 1
        return t


CAL_UGD     = Calendar(start_hour=8,  end_hour=16)
CAL_MATRONA = Calendar(start_hour=8,  end_hour=15)
CAL_LUNES   = Calendar(start_hour=8,  end_hour=16, weekdays=frozenset({0}))

# ========================
# Config dataclass
# ========================
@dataclass
class SimConfig:
    # Horizonte
    weeks_to_simulate:  int   = 52
    replications:       int   = 10
    random_seed_base:   int   = 202

    # ── NUEVO: control de paralelismo ─────────────────────────────────────────
    # 0 = automático (usa todos los CPUs disponibles)
    # N > 0 = usa exactamente N procesos
    n_workers: int = 0

    matrona_capacity: int = 1

    # Agenda semanal
    publish_lead_workdays: int = 5

    fixed_weekly_capacity:  int   =  16
    use_fixed_weekly_capacity: bool = False

    blocked_pct:    float = 0.32
    cnu_pct:        float = 0.00

    aps_reject_p: float = 0.05

    lead_short_hours: float = 48.0

    # Backlog inicial 1ra consulta
    backlog_low_n:  int = 2
    backlog_mid_n:  int = 138
    backlog_high_n: int = 248

    # Backlog POST inicial
    seed_post_backlog_at_t0:    bool  = True
    post_init_preq_n:           int   = 200
    post_init_control_n:        int   = 700
    post_init_surgery_waitlist_n: int = 170
    post_init_priority_share: dict = field(default_factory=lambda: {
        "high": 0.72, "mid": 0.26, "low": 0.02
    })
    post_backlog_beta_kappa: float = 10.0

    # Horario laboral
    work_start_h: int = 8
    work_end_h:   int = 16

    # Llegadas
    nb_arrivals_n: float = 8.99
    nb_arrivals_p: float = 0.7157

    # Reconversión
    reconv_capacity:           int   = 3
    reconv_duration_min:       float = 5.0
    reconv_sla_max_wait_workd: int   = 5

    # Agente
    agent_capacity:          int   = 1
    not_contactable_p: float = 0.15
    agent_limit_to_weekdays: bool  = True
    contact_attempt_min:     float = 5.0
    max_contact_attempts:    int   = 3
    case_time_long_min:      float = 5.0
    case_time_short_min:     float = 15.0
    weekly_cap_long:         int   = 20
    weekly_cap_short:        int   = 20
    confirm_ok_p:            float = 0.75
    confirm_decision_cancel_p: float = 0.22
    min_lead_publish_hours:  float = 0.0

    # 1ra consulta
    consult_duration_min:    float = 20.0
    gamma_shape:             float = 8.37
    gamma_scale:             float = 1.77
    day_share: dict = field(default_factory=lambda: {
        0: 0.09, 1: 0.41, 2: 0.19, 3: 0.15, 4: 0.16
    })
    unused_pct: float = 0.0

    # Post consulta
    endo_p:            float = 0.30
    endo_p_by_priority: bool = True
    endo_p_high:       float = 0.18
    endo_p_mid:        float = 0.12
    endo_p_low:        float = 0.08
    special_dx_share:  float = 0.20
    special_dx_booking_min: float = 5.0

    # Matrona
    matrona_booking_min:     float = 5.0
    matrona_review_min_low:  float = 20.0
    matrona_review_min_high: float = 30.0
    matrona_first_review_min: float = 5.0

    # Control post
    use_fixed_post_control_hours : bool = False
    fixed_post_control_capacity:  int  = 40
    post_control_hours_base:      float = 3.5
    post_control_hours_mult:      float = 26.8
    post_control_beta_a:          float = 0.945
    post_control_beta_b:          float = 1.63
    day_share_post: dict = field(default_factory=lambda: {
        0: 0.09, 1: 0.43, 2: 0.18, 3: 0.15, 4: 0.15
    })

    blocked_pct_post_control: float = 0.34
    nsp_matrona_p:    float = 0.09

    # Rutas post
    control_share: float = 0.67
    preq_share:    float = 0.05

    control_assign_matrona_p:  float = 0.70
    control_labs_p:            float = 0.67
    empty_control_p_ugd:       float = 0.3
    max_control_rebooks:       int   = 5
    max_empty_control_repeats: int   = 2

    # Ecografías y labs
    ugd_us_per_week:    int   = 25
    ugd_us_extra_pct:   float = 0.00
    ugd_us_duration_min: float = 60.0
    ugd_lab_per_week:   int   = 54
    ugd_lab_duration_min: float = 60.0
    ugd_lab_lead_days:  int   = 2

    mat_us_per_week:    int   = 25
    mat_us_duration_min: float = 60.0
    mat_us_lead_days:   int   = 2
    mat_lab_per_week:   int   = 25
    mat_lab_duration_min: float = 60.0
    mat_lab_lead_days:  int   = 2

    # Prequirúrgico
    preq_select_per_week:     int   = 6
    preq_special_treat_p:     float = 0.10
    preq_special_treat_days:  int   = 7

    # Quirúrgico
    quir_altered_p:                    float = 0.50
    quir_altered_option_specialist_p:  float = 0.50
    anesth_visit_min:      float = 20.0
    anesth_lead_days_min:  int   = 21
    anesth_lead_days_max:  int   = 30
    quir_altered_treatment_days: int = 7

    # Cirugía
    surg_throughput_per_week: int = 20
    crs_lead_days_min: int = 60
    crs_lead_days_max: int = 90

    # Post-op
    becado_visit_min: float = 20.0
    becado_request_after_surgery_days_min: int = 4
    becado_request_after_surgery_days_max: int = 6
    becado_book_within_days: int = 5

    @property
    def sim_time_min(self) -> float:
        return self.weeks_to_simulate * 7 * 24 * 60

    @property
    def days_to_simulate(self) -> int:
        return self.weeks_to_simulate * 7


CFG = SimConfig()

# ========================
# Distribuciones
# ========================
POST_BACKLOG_WAIT_STATS = {
    "high": {"min": 7,  "max": 544, "mean": 125.9667832},
    "mid":  {"min": 7,  "max": 684, "mean": 113.9665072},
    "low":  {"min": 14, "max": 643, "mean": 132.7058824},
}

def draw_post_backlog_wait_days(rng, priority: str, kappa: float = 10.0) -> float:
    s = POST_BACKLOG_WAIT_STATS[priority]
    a, b, mu = float(s["min"]), float(s["max"]), float(s["mean"])
    if b <= a:
        return a
    m = (mu - a) / (b - a)
    m = min(max(m, 1e-6), 1 - 1e-6)
    alpha = max(1e-3, m * kappa)
    beta  = max(1e-3, (1.0 - m) * kappa)
    return a + (b - a) * rng.beta(alpha, beta)

def draw_backlog_days_high(rng) -> float:
    return 7 + 274 * rng.beta(0.947, 1.37)

def draw_backlog_days_mid(rng) -> float:
    return 7 + 238 * rng.beta(0.747, 0.899)

def draw_daily_arrival_count(rng, cfg) -> int:
    count = 0
    while count <= 0:
        count = int(rng.negative_binomial(cfg.nb_arrivals_n, cfg.nb_arrivals_p))
    return count

def local_adjust_minutes_by_priority(rng, priority: str, cfg) -> float:
    workday_minutes = (cfg.work_end_h - cfg.work_start_h) * 60
    if priority == Priority.HIGH:
        return float(rng.uniform(0.5 * workday_minutes, 2.0 * workday_minutes))
    elif priority == Priority.MID:
        return float(rng.uniform(1.0 * workday_minutes, 4.0 * workday_minutes))
    else:
        return float(rng.uniform(1.0 * workday_minutes, 5.0 * workday_minutes))

# ========================
# Utilidades calendario
# ========================
def is_work_minute(t): return CAL_UGD.is_work_minute(t)
def next_work_minute(t): return CAL_UGD.next_work_minute(t)
def consume_work_minutes(env, minutes): yield from CAL_UGD.consume(env, minutes)
def total_work_minutes_until(h): return CAL_UGD.total_work_minutes_until(h)
def add_workdays(t0, wd): return CAL_UGD.add_workdays(t0, wd)
def workdays_before(t0, wd): return CAL_UGD.workdays_before(t0, wd)

def is_matrona_work_minute(t): return CAL_MATRONA.is_work_minute(t)
def next_matrona_work_minute(t): return CAL_MATRONA.next_work_minute(t)
def total_matrona_work_minutes_until(h): return CAL_MATRONA.total_work_minutes_until(h)

def next_agent_minute(t, agent_days, cfg):
    cal = Calendar(start_hour=cfg.work_start_h, end_hour=cfg.work_end_h,
                   weekdays=frozenset(agent_days))
    return cal.next_work_minute(t)

def total_agent_minutes_until(h, agent_days, cfg):
    cal = Calendar(start_hour=cfg.work_start_h, end_hour=cfg.work_end_h,
                   weekdays=frozenset(agent_days))
    return cal.total_work_minutes_until(h)

def week_index_from_time(t): return int(t // (7 * 24 * 60))

AGENT_BOOKING_WEEKDAYS_FIRST = {0,1,2,3,4,5}
AGENT_BOOKING_WEEKDAYS_POST  = {0}

# ========================
# Estructuras
# ========================
@dataclass(frozen=True)
class Patient:
    pid: int
    priority: str
    days_wait_at_start: float
    enqueued_at: float
    is_backlog: bool
    endometriosis: bool

@dataclass
class Booking:
    patient: Patient
    booked_at: float
    slot_time: float
    lead_hours: float
    result: Optional[str] = None

@dataclass
class PostNeed:
    nid: int
    patient: Patient
    not_before: float
    source: str
    kind: str
    requires_labs: bool
    requires_us: bool
    empty_repeat_left: int
    rebooks_left: int
    queue_priority: str
    done_event: simpy.Event

# ========================
# KPIs
# ========================
class KPIs:
    def __init__(self):
        self.bookings = 0
        self.bookings_by_pri = {p: 0 for p in Priority}
        self.attended = 0
        self.attended_by_pri = {p: 0 for p in Priority}
        self.not_contactable = 0
        self.unreachable_removed = 0
        self.canceled = 0
        self.cnu_count = 0

        self.total_time_all = []
        self.total_time_all_by_pri = {p: [] for p in Priority}
        self.first_closed_total_time_all = []
        self.first_closed_total_time_all_by_pri = {p: [] for p in Priority}
        self.first_closed_backlog_time_all = []
        self.first_closed_process_time_all = []
        self.first_attended_total_time_all = []
        self.first_attended_total_time_all_by_pri = {p: [] for p in Priority}
        self.first_attended_backlog_time_all = []
        self.first_attended_process_time_all = []
        self.pre_system_wait_by_pri = {p: [] for p in Priority}
        self.local_adjust_by_pri    = {p: [] for p in Priority}
        self.agendamiento_by_pri    = {p: [] for p in Priority}
        self.wait_before_care_by_pri= {p: [] for p in Priority}

        self.agent_time_total = 0.0
        self.slot_minutes_published  = 0.0
        self.specialist_minutes_used = 0.0

        self.ts_time_min = []
        self.ts_cum_wait_total = []
        self.ts_cum_att_total  = []
        self.ts_diff_total     = []

        self.post_entered = 0
        self.post_completed = 0
        self.post_entered_by_pri  = {p: 0 for p in Priority}
        self.post_completed_by_pri= {p: 0 for p in Priority}
        self.post_route_counts = {"control": 0, "preq": 0, "quir": 0}
        self.post_control_bookings = 0
        self.post_control_attended = 0
        self.post_control_blocked  = 0
        self.post_control_nsp      = 0
        self.post_control_empty    = 0
        self.post_control_bookings_by_source = {"matrona": 0, "ugd": 0}
        self.post_control_att_by_source      = {"matrona": 0, "ugd": 0}
        self.post_control_minutes_published  = 0.0
        self.post_control_minutes_used       = 0.0
        self.post_agent_time_total   = 0.0
        self.post_matrona_time_total = 0.0
        self.post_anesth_attended = 0
        self.post_becado_attended = 0

        self.post_total_time_all  = []
        self.total_time_full_all  = []

        self.ts_post_time_min = []
        self.ts_post_entered_cum  = []
        self.ts_post_completed_cum= []
        self.ts_post_in_process   = []
        self.ts_post_control_attended_cum = []

        self.waitq_minutes   = defaultdict(list)
        self.waitslot_minutes= defaultdict(list)

        self.ts_week_idx  = []
        self.ts_wl_first  = []
        self.ts_wl_control= []
        self.ts_wl_preq   = []
        self.ts_wl_quir   = []

        self.ts_month_idx   = []
        self.ts_m_wl_first  = []
        self.ts_m_wl_control= []
        self.ts_m_wl_preq   = []
        self.ts_m_wl_quir   = []

        self.quir_altered_total     = 0
        self.quir_altered_treatment = 0
        self.quir_altered_specialist= 0
        self.quir_altered_anesth    = 0

    def add_waitq(self, key, minutes):
        if minutes is not None:
            self.waitq_minutes[key].append(float(minutes))

    def add_waitslot(self, key, minutes):
        if minutes is not None:
            self.waitslot_minutes[key].append(float(minutes))


# ========================
# Modelo base (ABC)
# ========================
class ClinicModelBase(abc.ABC):
    @abc.abstractmethod
    def arrival_at(self, t_arr): ...
    @abc.abstractmethod
    def reconversion_pipeline(self, patient): ...
    @abc.abstractmethod
    def start_post_consulta(self, patient, t_first_end): ...
    @abc.abstractmethod
    def route_control(self, patient): ...
    @abc.abstractmethod
    def route_preq(self, patient): ...
    @abc.abstractmethod
    def route_quir(self, patient, from_preq): ...
    @abc.abstractmethod
    def route_surgery_waitlist(self, patient): ...
    @abc.abstractmethod
    def agent_dispatcher_first(self): ...
    @abc.abstractmethod
    def agent_dispatcher_post_ugd(self): ...
    @abc.abstractmethod
    def dispatcher_post_matrona(self): ...


# ========================
# Modelo concreto
# ========================
class ClinicModelAdjusted(ClinicModelBase):

    def __init__(self, env, rng, cfg: SimConfig = None):
        self.env = env
        self.rng = rng
        self.cfg = cfg or CFG
        self.kpis = KPIs()
        self._id_seq       = itertools.count(1)
        self._post_need_id = itertools.count(1)

        self.patient_state   = {}
        self.active_non_surgery = set()
        self.daily_transition_errors = []
        self.validation_checks = {}
        self.patient_trace = {}

        self.first_slot_lock = simpy.Resource(env, capacity=1)
        self.reconv = simpy.PriorityResource(env, capacity=self.cfg.reconv_capacity)
        self.agent  = simpy.Resource(env, capacity=self.cfg.agent_capacity)

        self.wait_high, self.wait_mid, self.wait_low = deque(), deque(), deque()
        self.contact_attempts    = defaultdict(int)
        self.ready_for_booking_at= {}
        self.requeue_not_before  = defaultdict(float)
        self.counted_in_cum_wait = set()
        self.cum_removed_total   = 0
        self.cum_removed_by_pri  = {p: 0 for p in Priority}

        self.slot_times = deque()
        self.week_caps_used = defaultdict(lambda: {"long": 0, "short": 0})
        self.published_at = None
        self.earliest_bookable_time = None

        back_total = self.cfg.backlog_low_n + self.cfg.backlog_mid_n + self.cfg.backlog_high_n
        self.cum_wait_total = back_total
        self.att_total_cum  = 0

        self.kpis.ts_time_min.append(0.0)
        self.kpis.ts_cum_wait_total.append(self.cum_wait_total)
        self.kpis.ts_cum_att_total.append(0)
        self.kpis.ts_diff_total.append(self.cum_wait_total)

        self.matrona      = simpy.Resource(env, capacity=self.cfg.matrona_capacity)
        self.post_slot_lock = simpy.PriorityResource(env, capacity=1)
        self.anesthesist  = simpy.Resource(env, capacity=1)
        self.becado       = simpy.Resource(env, capacity=1)

        self.post_control_slots = deque()

        self.post_wait_high_matrona = deque()
        self.post_wait_mid_matrona  = deque()
        self.post_wait_low_matrona  = deque()
        self.post_wait_high_ugd     = deque()
        self.post_wait_mid_ugd      = deque()
        self.post_wait_low_ugd      = deque()

        self.ugd_us_slots  = deque()
        self.ugd_us_lock   = simpy.Resource(env, capacity=1)
        self.ugd_lab_slots = deque()
        self.ugd_lab_lock  = simpy.Resource(env, capacity=1)
        self.mat_us_slots  = deque()
        self.mat_lab_slots = deque()
        self.mat_exam_lock = simpy.Resource(env, capacity=1)

        self.preq_queue         = deque()
        self.preq_enq_time      = {}
        self.preq_select_events = {}

        self.surg_queue_len    = 0
        self.surgery_tokens    = simpy.Container(env, init=0, capacity=10**9)
        self.surgery_waiting_open = 0

        self.post_in_process   = set()
        self.post_start_time   = {}
        self.post_entered_cum  = 0
        self.post_completed_cum= 0
        self.post_booked_open  = defaultdict(int)

        env.process(self.weekly_pre_publish_loop())
        env.process(self.weekly_post_publish_loop())
        env.process(self.weekly_ugd_us_slots_loop())
        env.process(self.weekly_ugd_lab_slots_loop())
        env.process(self.weekly_mat_exam_slots_loop())
        env.process(self.preq_weekly_selector_loop())
        env.process(self.weekly_surgery_tokens_loop())

        self._seed_backlog()
        self._seed_post_backlog()
        self.push_post_timeseries_point()

        env.process(self.daily_arrivals())
        env.process(self.agent_dispatcher_first())
        env.process(self.agent_dispatcher_post_ugd())
        env.process(self.dispatcher_post_matrona())
        env.process(self.daily_monitor_end_of_day())
        env.process(self.record_queue_snapshot_t0())
        env.process(self.weekly_queue_monitor())

    # ── Helpers internos ────────────────────────────────────────────────────
    def _set_patient_state(self, pid, state):
        self.patient_state[pid] = state

    def _mark_active_non_surgery(self, pid, active=True):
        if pid is None:
            return
        if active:
            self.active_non_surgery.add(pid)
        else:
            self.active_non_surgery.discard(pid)

    def _register_patient(self, p: Patient):
        if p.pid not in self.patient_trace:
            self.patient_trace[p.pid] = {
                "pid": p.pid,
                "priority": p.priority,
                "endometriosis": bool(p.endometriosis),
                "t_enqueue": float(p.enqueued_at),
                "first_done_at": None,
                "first_closed_at": None,
                "first_attended_at": None,
                "first_closure_reason": None,
                "post_done_at": None,
                "post_route": None,
                "waitq": defaultdict(float),
                "waitslot": defaultdict(float),
            }

    def _trace_add_waitq(self, pid, key: str, minutes: float):
        self.kpis.add_waitq(key, minutes)
        if pid is None:
            return
        tr = self.patient_trace.get(pid)
        if tr:
            tr["waitq"][key] += float(minutes)

    def _trace_add_waitslot(self, pid, key: str, minutes: float):
        self.kpis.add_waitslot(key, minutes)
        if pid is None:
            return
        tr = self.patient_trace.get(pid)
        if tr:
            tr["waitslot"][key] += float(minutes)

    def _record_first_closure(self, patient, total_minutes, reason, attended=False):
        pri = patient.priority
        total_minutes  = float(total_minutes)
        backlog_minutes= float(max(0.0, -float(patient.enqueued_at)))
        process_minutes= float(max(0.0, total_minutes - backlog_minutes))

        self.kpis.total_time_all.append(total_minutes)
        self.kpis.total_time_all_by_pri[pri].append(total_minutes)
        self.kpis.first_closed_total_time_all.append(total_minutes)
        self.kpis.first_closed_total_time_all_by_pri[pri].append(total_minutes)
        self.kpis.first_closed_backlog_time_all.append(backlog_minutes)
        self.kpis.first_closed_process_time_all.append(process_minutes)
        if attended:
            self.kpis.first_attended_total_time_all.append(total_minutes)
            self.kpis.first_attended_total_time_all_by_pri[pri].append(total_minutes)
            self.kpis.first_attended_backlog_time_all.append(backlog_minutes)
            self.kpis.first_attended_process_time_all.append(process_minutes)

        tr = self.patient_trace.get(patient.pid)
        if tr:
            tr["first_closed_at"] = float(self.env.now)
            tr["first_closure_reason"] = str(reason)
            if attended:
                tr["first_attended_at"] = float(self.env.now)

    def _resolve_done_event(self, need, outcome):
        if not need.done_event.triggered:
            need.done_event.succeed(outcome)

    def _calendar_end_today(self, t, calendar, agent_days=None):
        day = int(t // (24 * 60))
        if calendar == 'work':
            return day * 24 * 60 + self.cfg.work_end_h * 60
        if calendar == 'matrona':
            return day * 24 * 60 + 15 * 60
        if calendar == 'agent':
            return day * 24 * 60 + self.cfg.work_end_h * 60
        raise ValueError(f"Calendario no soportado: {calendar}")

    def _next_with_capacity(self, t, minutes, calendar, agent_days=None):
        while True:
            if calendar == 'work':
                t0 = CAL_UGD.next_work_minute(t)
            elif calendar == 'matrona':
                t0 = CAL_MATRONA.next_work_minute(t)
            elif calendar == 'agent':
                cal = Calendar(start_hour=self.cfg.work_start_h,
                               end_hour=self.cfg.work_end_h,
                               weekdays=frozenset(agent_days or AGENT_BOOKING_WEEKDAYS_FIRST))
                t0 = cal.next_work_minute(t)
            else:
                raise ValueError(f"Calendario no soportado: {calendar}")
            if self._calendar_end_today(t0, calendar, agent_days) - t0 >= minutes - 1e-9:
                return t0
            t = self._calendar_end_today(t0, calendar, agent_days) + 1

    def _service_loop(self, resource, minutes, calendar, pid=None, waitq_key=None, agent_days=None):
        while True:
            t_ready = self._next_with_capacity(self.env.now, minutes, calendar, agent_days)
            if t_ready > self.env.now:
                yield self.env.timeout(t_ready - self.env.now)
            t_q = self.env.now
            with resource.request() as req:
                yield req
                if self._calendar_end_today(self.env.now, calendar, agent_days) - self.env.now >= minutes - 1e-9:
                    if waitq_key:
                        self._trace_add_waitq(pid, waitq_key, self.env.now - t_q)
                    self._mark_active_non_surgery(pid, True)
                    try:
                        yield self.env.timeout(minutes)
                    finally:
                        self._mark_active_non_surgery(pid, False)
                    return
            t_next = self._next_with_capacity(self.env.now + 1, minutes, calendar, agent_days)
            if t_next > self.env.now:
                yield self.env.timeout(t_next - self.env.now)

    def _matrona_fixed(self, minutes, pid=None):
        yield self.env.process(self._service_loop(
            self.matrona, minutes, 'matrona', pid=pid, waitq_key='matrona'))
        self.kpis.post_matrona_time_total += minutes

    def _matrona_work(self, lo, hi, pid=None):
        minutes = float(self.rng.uniform(lo, hi))
        yield from self._matrona_fixed(minutes, pid=pid)

    def _book_from_pool(self, pool, lock_res, not_before, waitslot_key, pid=None):
        t_request = self.env.now
        retries = 0
        max_retries = max(1, int(self.cfg.sim_time_min // 10))
        while True:
            if self.env.now >= self.cfg.sim_time_min or retries > max_retries:
                log.warning("Pool agotado pid=%s key=%s t=%.0f", pid, waitslot_key, self.env.now)
                return None
            t0 = max(not_before, self.env.now)
            t_q = self.env.now
            with lock_res.request() as req:
                yield req
                self._trace_add_waitq(pid, 'pool_lock', self.env.now - t_q)
                while pool and pool[0] <= self.env.now:
                    pool.popleft()
                idx = None
                for i, s in enumerate(pool):
                    if s >= t0:
                        idx = i
                        break
                if idx is not None:
                    slot_t = pool[idx]
                    del pool[idx]
                    self._trace_add_waitslot(pid, waitslot_key, slot_t - t_request)
                    return slot_t
            retries += 1
            yield self.env.timeout(10)

    def _endo_p(self, priority: str) -> float:
        if not self.cfg.endo_p_by_priority:
            return self.cfg.endo_p
        if priority == Priority.HIGH: return self.cfg.endo_p_high
        if priority == Priority.MID:  return self.cfg.endo_p_mid
        return self.cfg.endo_p_low

    def _make_patient(self, priority, days_wait, is_backlog):
        pid  = next(self._id_seq)
        enq  = self.env.now - days_wait * 24 * 60
        endo = (self.rng.random() < self._endo_p(priority))
        p    = Patient(pid=pid, priority=priority, days_wait_at_start=days_wait,
                       enqueued_at=enq, is_backlog=is_backlog, endometriosis=endo)
        self.kpis.pre_system_wait_by_pri[priority].append(max(0.0, -enq))
        if is_backlog:
            self.counted_in_cum_wait.add(pid)
        self._register_patient(p)
        self._set_patient_state(pid, 'first_backlog' if is_backlog else 'first_new')
        return p

    def _seed_backlog(self):
        for _ in range(self.cfg.backlog_low_n):
            d = self.rng.uniform(42, 242)
            self.env.process(self.reconversion_pipeline(
                self._make_patient(Priority.LOW, d, True)))
        for _ in range(self.cfg.backlog_mid_n):
            d = draw_backlog_days_mid(self.rng)
            self.env.process(self.reconversion_pipeline(
                self._make_patient(Priority.MID, d, True)))
        for _ in range(self.cfg.backlog_high_n):
            d = draw_backlog_days_high(self.rng)
            self.env.process(self.reconversion_pipeline(
                self._make_patient(Priority.HIGH, d, True)))

    def _seed_post_backlog(self):
        if not self.cfg.seed_post_backlog_at_t0:
            return
        sh = self.cfg.post_init_priority_share
        ph = float(sh.get("high", 0.72))
        pm = float(sh.get("mid",  0.26))
        pl = float(sh.get("low",  max(0.0, 1.0 - ph - pm)))
        tot = ph + pm + pl
        if tot <= 0:
            ph, pm, pl = 0.72, 0.26, 0.02
        ph, pm, pl = ph / tot, pm / tot, pl / tot

        def draw_pr():
            r = self.rng.random()
            if r < ph:     return Priority.HIGH
            if r < ph+pm:  return Priority.MID
            return Priority.LOW

        def new_post_patient(pr):
            d   = draw_post_backlog_wait_days(self.rng, pr, self.cfg.post_backlog_beta_kappa)
            enq = self.env.now - d * 24 * 60
            p   = Patient(pid=next(self._id_seq), priority=pr, days_wait_at_start=d,
                          enqueued_at=enq, is_backlog=True,
                          endometriosis=(self.rng.random() < self._endo_p(pr)))
            self._register_patient(p)
            return p

        for _ in range(int(self.cfg.post_init_preq_n)):
            self._start_post_backlog(new_post_patient(draw_pr()), "preq")
        for _ in range(int(self.cfg.post_init_control_n)):
            self._start_post_backlog(new_post_patient(draw_pr()), "control")
        for _ in range(int(self.cfg.post_init_surgery_waitlist_n)):
            self._start_post_backlog(new_post_patient(draw_pr()), "quir")

    def _start_post_backlog(self, patient: Patient, route: str):
        t_first_end = 0.0
        if t_first_end >= self.cfg.sim_time_min:
            return
        self.kpis.post_entered += 1
        self.kpis.post_entered_by_pri[patient.priority] += 1
        self.post_entered_cum += 1
        self.post_in_process.add(patient.pid)
        self.post_start_time[patient.pid] = t_first_end
        tr = self.patient_trace.get(patient.pid)
        if tr:
            tr['first_done_at'] = float(t_first_end)

        route_map = {
            "control": self.route_control,
            "preq":    self.route_preq,
            "quir":    self.route_surgery_waitlist,
        }
        fn = route_map.get(route, self.route_control)
        self.kpis.post_route_counts[route if route in ("control","preq","quir") else "control"] += 1
        if tr:
            tr['post_route'] = route
        self.env.process(fn(patient))

    def daily_arrivals(self):
        day = 0
        while True:
            if self.env.now >= self.cfg.sim_time_min:
                break
            next_day = (day + 1) * 24 * 60
            count = draw_daily_arrival_count(self.rng, self.cfg)
            for _ in range(count):
                t_arr = day * 24 * 60 + self.rng.random() * (24 * 60)
                if t_arr < self.cfg.sim_time_min:
                    self.env.process(self.arrival_at(t_arr))
            yield self.env.timeout(next_day - self.env.now)
            day += 1

    def arrival_at(self, t_arr):
        if self.env.now < t_arr:
            yield self.env.timeout(t_arr - self.env.now)

        if self.rng.random() < self.cfg.aps_reject_p:
            log.debug("Paciente rechazado APS en t=%.0f", self.env.now)
            return

        r = self.rng.random()
        if r < 0.00005:
            pr = Priority.LOW
        elif r < (0.00005 + 0.35):
            pr = Priority.MID
        else:
            pr = Priority.HIGH

        endo = (self.rng.random() < self._endo_p(pr))
        p = Patient(pid=next(self._id_seq), priority=pr, days_wait_at_start=0.0,
                    enqueued_at=self.env.now, is_backlog=False, endometriosis=endo)
        self.kpis.pre_system_wait_by_pri[pr].append(0.0)
        self._register_patient(p)
        self._set_patient_state(p.pid, 'first_arrival')
        self.env.process(self.reconversion_pipeline(p))

    def reconversion_pipeline(self, patient):
        sla_deadline = CAL_UGD.add_workdays(self.env.now, self.cfg.reconv_sla_max_wait_workd)
        prio_map = {Priority.HIGH: 0, Priority.MID: 1, Priority.LOW: 2}
        self._set_patient_state(patient.pid, 'waiting_reconversion')

        while True:
            t_ready = self._next_with_capacity(self.env.now, self.cfg.reconv_duration_min, 'work')
            if t_ready > self.env.now:
                yield self.env.timeout(t_ready - self.env.now)
            t_start_wait = self.env.now
            req = self.reconv.request(priority=prio_map[patient.priority])
            got = False
            while self.env.now < sla_deadline:
                res = yield req | self.env.timeout(1)
                if req in res:
                    got = True
                    break
            if not got:
                try:
                    self.reconv.release(req)
                except Exception:
                    pass
                req = self.reconv.request(priority=-1)
                yield req
            if self._calendar_end_today(self.env.now, 'work') - self.env.now >= self.cfg.reconv_duration_min - 1e-9:
                self._trace_add_waitq(patient.pid, 'reconv', self.env.now - t_start_wait)
                self._set_patient_state(patient.pid, 'in_reconversion')
                self._mark_active_non_surgery(patient.pid, True)
                try:
                    yield self.env.timeout(self.cfg.reconv_duration_min)
                finally:
                    self._mark_active_non_surgery(patient.pid, False)
                self.reconv.release(req)
                break
            self.reconv.release(req)

        self._set_patient_state(patient.pid, 'waiting_matrona_first')
        yield self.env.process(self._service_loop(
            self.matrona, self.cfg.matrona_first_review_min, 'matrona',
            pid=patient.pid, waitq_key='matrona'))
        self.kpis.post_matrona_time_total += self.cfg.matrona_first_review_min

        adj = local_adjust_minutes_by_priority(self.rng, patient.priority, self.cfg)
        if adj > 0:
            self._set_patient_state(patient.pid, 'local_adjust_wait')
            yield self.env.timeout(adj)

        tr = self.patient_trace.get(patient.pid)
        if tr and ('special_dx_route' not in tr):
            tr['special_dx_route'] = bool(self.rng.random() < self.cfg.special_dx_share)

        if tr and tr.get('special_dx_route', False):
            self._set_patient_state(patient.pid, 'special_prefirst_route')
            self.env.process(self.special_prefirst_route(patient))
            return

        self.ready_for_booking_at[patient.pid] = self.env.now
        self._set_patient_state(patient.pid, 'waiting_first_booking')
        if patient.priority == Priority.HIGH:
            self.wait_high.append(patient)
        elif patient.priority == Priority.MID:
            self.wait_mid.append(patient)
        else:
            self.wait_low.append(patient)

        if patient.pid not in self.counted_in_cum_wait:
            self.counted_in_cum_wait.add(patient.pid)
            self.cum_wait_total += 1

    def _book_first_consult_slot(self, not_before, pid=None):
        t_request = self.env.now
        while True:
            if self.env.now >= self.cfg.sim_time_min:
                return None
            t0 = max(not_before, self.env.now, self.earliest_bookable_time or 0.0)
            with self.first_slot_lock.request() as req:
                yield req
                while self.slot_times and self.slot_times[0] <= self.env.now:
                    self.slot_times.popleft()
                chosen_idx = None
                slot_t = wk = is_long = None
                for i, s in enumerate(self.slot_times):
                    if s < t0:
                        continue
                    lead_hours = max(0.0, (s - self.env.now) / 60.0)
                    is_long = (lead_hours >= self.cfg.lead_short_hours)
                    wk      = week_index_from_time(s)
                    used    = self.week_caps_used[wk]['long' if is_long else 'short']
                    lim     = self.cfg.weekly_cap_long if is_long else self.cfg.weekly_cap_short
                    if used < lim:
                        chosen_idx = i
                        slot_t = s
                        break
                if chosen_idx is not None:
                    del self.slot_times[chosen_idx]
                    self.week_caps_used[wk]['long' if is_long else 'short'] += 1
                    self._trace_add_waitslot(pid, 'first_appointment_lead', slot_t - t_request)
                    return slot_t
            yield self.env.timeout(10)

    def special_prefirst_route(self, patient):
        tr = self.patient_trace.get(patient.pid)
        if tr:
            tr['special_dx_route'] = True
            tr['special_dx_started_at'] = float(self.env.now)

        def _fail(reason):
            self._set_patient_state(patient.pid, reason)
            tts = self.env.now - patient.enqueued_at
            self._record_first_closure(patient, tts, reason=reason, attended=False)
            self.cum_removed_total += 1
            self.cum_removed_by_pri[patient.priority] += 1

        yield self.env.process(self._matrona_fixed(self.cfg.special_dx_booking_min, pid=patient.pid))
        us_t = yield self.env.process(self._book_from_pool(
            self.mat_us_slots, self.mat_exam_lock,
            not_before=self.env.now + self.cfg.mat_us_lead_days * 24 * 60,
            waitslot_key='mat_us_lead', pid=patient.pid))
        if us_t is None:
            _fail('special_prefirst_failed_us')
            return

        yield self.env.process(self._matrona_fixed(self.cfg.special_dx_booking_min, pid=patient.pid))
        lab_t = yield self.env.process(self._book_from_pool(
            self.mat_lab_slots, self.mat_exam_lock,
            not_before=self.env.now + self.cfg.mat_lab_lead_days * 24 * 60,
            waitslot_key='mat_lab_lead', pid=patient.pid))
        if lab_t is None:
            _fail('special_prefirst_failed_lab')
            return

        exams_ready = max(us_t + self.cfg.mat_us_duration_min,
                          lab_t + self.cfg.mat_lab_duration_min)
        if self.env.now < exams_ready:
            yield self.env.timeout(exams_ready - self.env.now)

        yield self.env.process(self._matrona_fixed(self.cfg.special_dx_booking_min, pid=patient.pid))
        slot_t = yield self.env.process(self._book_first_consult_slot(exams_ready, pid=patient.pid))
        if slot_t is None:
            _fail('special_prefirst_failed_first_slot')
            return

        ready_t = self.ready_for_booking_at.get(patient.pid, patient.enqueued_at)
        self.kpis.agendamiento_by_pri[patient.priority].append(self.env.now - ready_t)
        self.kpis.bookings += 1
        self.kpis.bookings_by_pri[patient.priority] += 1
        booking = Booking(patient=patient, booked_at=self.env.now, slot_time=slot_t,
                          lead_hours=max(0.0, (slot_t - self.env.now) / 60.0))
        self._set_patient_state(patient.pid, 'booked_first_consult_special_dx')
        if tr:
            tr['special_dx_first_booked_at'] = float(self.env.now)
            tr['special_dx_exams_ready_at']  = float(exams_ready)
        self.env.process(self.appointment_event_first(booking))
        if booking.lead_hours >= self.cfg.lead_short_hours:
            self.env.process(self.confirmation_lead_first(booking))

    def pick_patient_first(self):
        now = self.env.now
        def eligible(q):
            return [p for p in q if self.requeue_not_before.get(p.pid, 0.0) <= now]
        for q in (self.wait_high, self.wait_mid, self.wait_low):
            cand = eligible(q)
            if cand:
                return max(cand, key=lambda p: (now - p.enqueued_at))
        return None

    def remove_patient_first(self, p: Patient):
        q = (self.wait_high if p.priority == Priority.HIGH
             else self.wait_mid if p.priority == Priority.MID
             else self.wait_low)
        try:
            q.remove(p)
        except ValueError:
            pass

    def agent_dispatcher_first(self):
        while True:
            if self.env.now >= self.cfg.sim_time_min:
                break
            t0 = (next_agent_minute(self.env.now, AGENT_BOOKING_WEEKDAYS_FIRST, self.cfg)
                  if self.cfg.agent_limit_to_weekdays
                  else CAL_UGD.next_work_minute(self.env.now))
            if t0 > self.env.now:
                yield self.env.timeout(t0 - self.env.now)

            while self.slot_times and self.slot_times[0] <= self.env.now:
                self.slot_times.popleft()
            if not self.slot_times:
                yield self.env.timeout(5)
                continue

            patient = self.pick_patient_first()
            if patient is None:
                yield self.env.timeout(5)
                continue

            t_ready = self._next_with_capacity(
                self.env.now,
                self.cfg.contact_attempt_min + self.cfg.case_time_short_min,
                'agent', AGENT_BOOKING_WEEKDAYS_FIRST)
            if t_ready > self.env.now:
                yield self.env.timeout(t_ready - self.env.now)

            t_q = self.env.now
            with self.agent.request() as req:
                yield req
                if (self._calendar_end_today(self.env.now, 'agent', AGENT_BOOKING_WEEKDAYS_FIRST)
                        - self.env.now < (self.cfg.contact_attempt_min + self.cfg.case_time_short_min - 1e-9)):
                    continue
                self._trace_add_waitq(patient.pid, 'agent_first', self.env.now - t_q)
                self._set_patient_state(patient.pid, 'contacting_first')
                self._mark_active_non_surgery(patient.pid, True)
                try:
                    yield self.env.timeout(self.cfg.contact_attempt_min)
                finally:
                    self._mark_active_non_surgery(patient.pid, False)
                self.kpis.agent_time_total += self.cfg.contact_attempt_min

                if self.rng.random() < self.cfg.not_contactable_p:
                    self.kpis.not_contactable += 1
                    self.contact_attempts[patient.pid] += 1
                    if self.contact_attempts[patient.pid] >= self.cfg.max_contact_attempts:
                        self.remove_patient_first(patient)
                        self.kpis.unreachable_removed += 1
                        self.cum_removed_total += 1
                        self.cum_removed_by_pri[patient.priority] += 1
                        self._set_patient_state(patient.pid, 'removed_unreachable_first')
                        tts = self.env.now - patient.enqueued_at
                        self._record_first_closure(patient, tts,
                                                   reason='removed_unreachable', attended=False)
                    else:
                        self._set_patient_state(patient.pid, 'waiting_first_booking')
                    continue

                earliest_ok = max(self.earliest_bookable_time or 0.0, self.env.now)
                chosen_idx = None
                for i, s in enumerate(self.slot_times):
                    if s >= earliest_ok:
                        chosen_idx = i
                        break
                if chosen_idx is None:
                    yield self.env.timeout(10)
                    continue

                slot_t     = self.slot_times[chosen_idx]
                lead_hours = max(0.0, (slot_t - self.env.now) / 60.0)
                is_long    = (lead_hours >= self.cfg.lead_short_hours)
                wk         = week_index_from_time(slot_t)
                used       = self.week_caps_used[wk]['long' if is_long else 'short']
                lim        = self.cfg.weekly_cap_long if is_long else self.cfg.weekly_cap_short
                if used >= lim:
                    yield self.env.timeout(10)
                    continue

                call_time = self.cfg.case_time_short_min if not is_long else self.cfg.case_time_long_min
                self._set_patient_state(patient.pid, 'booking_first')
                self._mark_active_non_surgery(patient.pid, True)
                try:
                    yield self.env.timeout(call_time)
                finally:
                    self._mark_active_non_surgery(patient.pid, False)
                self.kpis.agent_time_total += call_time

            while self.slot_times and self.slot_times[0] <= self.env.now:
                self.slot_times.popleft()
            if chosen_idx >= len(self.slot_times):
                continue

            slot_t = self.slot_times[chosen_idx]
            del self.slot_times[chosen_idx]
            self.remove_patient_first(patient)
            self.week_caps_used[wk]['long' if is_long else 'short'] += 1

            ready_t = self.ready_for_booking_at.get(patient.pid, patient.enqueued_at)
            self.kpis.agendamiento_by_pri[patient.priority].append(self.env.now - ready_t)
            self.kpis.bookings += 1
            self.kpis.bookings_by_pri[patient.priority] += 1
            self._trace_add_waitslot(patient.pid, 'first_appointment_lead', slot_t - self.env.now)
            booking = Booking(patient=patient, booked_at=self.env.now,
                              slot_time=slot_t, lead_hours=lead_hours)
            self._set_patient_state(patient.pid, 'booked_first_consult')
            self.env.process(self.appointment_event_first(booking))
            if lead_hours >= self.cfg.lead_short_hours:
                self.env.process(self.confirmation_lead_first(booking))

    def confirmation_lead_first(self, booking: Booking):
        t_confirm = booking.slot_time - self.cfg.lead_short_hours * 60
        if self.env.now < t_confirm:
            yield self.env.timeout(t_confirm - self.env.now)
        if booking.result is not None:
            return

        if self.rng.random() < self.cfg.confirm_decision_cancel_p:
            booking.result = 'cancel'
            if self.env.now < booking.slot_time:
                self.slot_times.appendleft(booking.slot_time)
            self.kpis.canceled += 1
            self._set_patient_state(booking.patient.pid, 'first_canceled')
            tts = self.env.now - booking.patient.enqueued_at
            self._record_first_closure(booking.patient, tts, reason='canceled', attended=False)
            return

        if self.rng.random() > self.cfg.confirm_ok_p:
            p = booking.patient
            self.requeue_not_before[p.pid] = self.env.now
            self.ready_for_booking_at[p.pid] = self.env.now
            if p.priority == Priority.HIGH:   self.wait_high.append(p)
            elif p.priority == Priority.MID:  self.wait_mid.append(p)
            else:                             self.wait_low.append(p)
            if self.env.now < booking.slot_time:
                self.slot_times.appendleft(booking.slot_time)
            booking.result = 'rebook'
            self._set_patient_state(p.pid, 'waiting_first_booking')

    def appointment_event_first(self, booking: Booking):
        if self.env.now < booking.slot_time:
            yield self.env.timeout(booking.slot_time - self.env.now)
        if booking.result in ('rebook', 'cancel'):
            return
        if self.rng.random() < self.cfg.unused_pct:
            self._set_patient_state(booking.patient.pid, 'first_unused_slot')
            return

        r_slot = self.rng.random()
        if r_slot < self.cfg.cnu_pct:
            self.kpis.cnu_count += 1
            p = booking.patient
            self.requeue_not_before[p.pid] = self.next_publish_after_first(self.env.now)
            self.ready_for_booking_at[p.pid] = self.env.now
            if p.priority == Priority.HIGH:   self.wait_high.append(p)
            elif p.priority == Priority.MID:  self.wait_mid.append(p)
            else:                             self.wait_low.append(p)
            self._set_patient_state(p.pid, 'waiting_first_booking')
            return

        if self.rng.random() < self.cfg.blocked_pct:
            p = booking.patient
            self.requeue_not_before[p.pid] = self.next_publish_after_first(self.env.now)
            self.ready_for_booking_at[p.pid] = self.env.now
            if p.priority == Priority.HIGH:   self.wait_high.append(p)
            elif p.priority == Priority.MID:  self.wait_mid.append(p)
            else:                             self.wait_low.append(p)
            self._set_patient_state(p.pid, 'waiting_first_booking')
            return

        pri = booking.patient.priority
        self.kpis.wait_before_care_by_pri[pri].append(
            booking.slot_time - booking.patient.enqueued_at)

        day = int(self.env.now // (24 * 60))
        end_of_day = day * 24 * 60 + self.cfg.work_end_h * 60
        if end_of_day - self.env.now < self.cfg.consult_duration_min - 1e-9:
            p = booking.patient
            self.requeue_not_before[p.pid] = self.next_publish_after_first(self.env.now)
            self.ready_for_booking_at[p.pid] = self.env.now
            if p.priority == Priority.HIGH:   self.wait_high.append(p)
            elif p.priority == Priority.MID:  self.wait_mid.append(p)
            else:                             self.wait_low.append(p)
            self._set_patient_state(p.pid, 'waiting_first_booking')
            return

        self._set_patient_state(booking.patient.pid, 'in_first_consult')
        self._mark_active_non_surgery(booking.patient.pid, True)
        try:
            yield self.env.timeout(self.cfg.consult_duration_min)
        finally:
            self._mark_active_non_surgery(booking.patient.pid, False)

        self.kpis.attended += 1
        self.kpis.attended_by_pri[pri] += 1
        self.kpis.specialist_minutes_used += self.cfg.consult_duration_min
        total_time = self.env.now - booking.patient.enqueued_at
        self._record_first_closure(booking.patient, total_time,
                                   reason='attended', attended=True)
        self.att_total_cum += 1
        self._set_patient_state(booking.patient.pid, 'post_started')
        self.start_post_consulta(booking.patient, t_first_end=self.env.now)

    def next_publish_after_first(self, t):
        w = week_index_from_time(t) + 1
        monday_8am = w * 7 * 24 * 60 + self.cfg.work_start_h * 60
        pub_t = CAL_UGD.workdays_before(monday_8am, self.cfg.publish_lead_workdays)
        return CAL_UGD.next_work_minute(max(pub_t, t))

    def start_post_consulta(self, patient, t_first_end):
        tr = self.patient_trace.get(patient.pid, {})
        if t_first_end >= self.cfg.sim_time_min:
            return
        self.kpis.post_entered += 1
        self.kpis.post_entered_by_pri[patient.priority] += 1
        self.post_entered_cum += 1
        self.post_in_process.add(patient.pid)
        self.post_start_time[patient.pid] = t_first_end
        tr2 = self.patient_trace.get(patient.pid)
        if tr2:
            tr2['first_done_at'] = float(t_first_end)

        if tr.get('special_dx_route', False):
            self.kpis.post_route_counts['quir'] += 1
            if tr2:
                tr2['post_route'] = 'quir'
            self._set_patient_state(patient.pid, 'waiting_surgery')
            self.env.process(self.route_quir(patient, from_preq=False))
            return

        r = self.rng.random()
        if r < self.cfg.control_share:
            self.kpis.post_route_counts['control'] += 1
            if tr2:
                tr2['post_route'] = 'control'
            self._set_patient_state(patient.pid, 'post_control_route')
            self.env.process(self.route_control(patient))
        elif r < self.cfg.control_share + self.cfg.preq_share:
            self.kpis.post_route_counts['preq'] += 1
            if tr2:
                tr2['post_route'] = 'preq'
            self._set_patient_state(patient.pid, 'post_preq_route')
            self.env.process(self.route_preq(patient))
        else:
            self.kpis.post_route_counts['quir'] += 1
            if tr2:
                tr2['post_route'] = 'quir'
            self._set_patient_state(patient.pid, 'waiting_surgery')
            self.env.process(self.route_quir(patient, from_preq=False))

    def route_control(self, patient):
        requires_labs = (self.rng.random() < self.cfg.control_labs_p)
        source = ('matrona' if self.rng.random() < self.cfg.control_assign_matrona_p
                  else 'ugd')

        if source == 'matrona':
            us_t = yield self.env.process(self._book_from_pool(
                self.mat_us_slots, self.mat_exam_lock,
                not_before=self.env.now + self.cfg.mat_us_lead_days * 24 * 60,
                waitslot_key='mat_us_lead', pid=patient.pid))
            if us_t is None:
                self._complete_post(patient)
                return
            us_end = us_t + self.cfg.mat_us_duration_min

            lab_end = None
            if requires_labs:
                lab_t = yield self.env.process(self._book_from_pool(
                    self.mat_lab_slots, self.mat_exam_lock,
                    not_before=self.env.now + self.cfg.mat_lab_lead_days * 24 * 60,
                    waitslot_key='mat_lab_lead', pid=patient.pid))
                if lab_t is None:
                    self._complete_post(patient)
                    return
                lab_end = lab_t + self.cfg.mat_lab_duration_min

            ready_for_control = (max(us_end, lab_end)
                                 if (requires_labs and lab_end is not None)
                                 else us_end)

        else:
            us_t = yield self.env.process(self._book_from_pool(
                self.ugd_us_slots, self.ugd_us_lock,
                not_before=self.env.now,
                waitslot_key='ugd_us_lead', pid=patient.pid))
            if us_t is None:
                self._complete_post(patient)
                return
            us_end = us_t + self.cfg.ugd_us_duration_min
            ready_for_control = us_end

        done = self.env.event()
        need = PostNeed(
            nid=next(self._post_need_id), patient=patient,
            not_before=ready_for_control, source=source, kind='control',
            requires_labs=requires_labs, requires_us=True,
            empty_repeat_left=(self.cfg.max_empty_control_repeats
                               if (source == 'ugd' and requires_labs) else 0),
            rebooks_left=self.cfg.max_control_rebooks,
            queue_priority=patient.priority, done_event=done,
        )
        self._enqueue_need(need)
        self._set_patient_state(patient.pid, 'waiting_post_need_control')
        outcome = yield done
        if outcome != 'attended':
            self._complete_post(patient)
            return
        yield self.env.process(self._post_discharge_or_monthly(patient))

    def route_preq(self, patient):
        ev = self.env.event()
        self.preq_select_events[patient.pid] = ev
        self.preq_enq_time[patient.pid] = (
            (self.env.now - patient.days_wait_at_start * 24 * 60)
            if (patient.is_backlog and patient.days_wait_at_start > 0)
            else self.env.now
        )
        self.preq_queue.append(patient)
        self._set_patient_state(patient.pid, 'waiting_preq_selection')
        yield ev

        tr = self.patient_trace.get(patient.pid)
        if tr:
            tr['post_route'] = 'preq'
            tr['entered_preq_at'] = float(self.env.now)

        yield self.env.process(self._matrona_work(
            self.cfg.matrona_review_min_low,
            self.cfg.matrona_review_min_high,
            pid=patient.pid))
        us_t = yield self.env.process(self._book_from_pool(
            self.mat_us_slots, self.mat_exam_lock,
            not_before=self.env.now + self.cfg.mat_us_lead_days * 24 * 60,
            waitslot_key='mat_us_lead', pid=patient.pid))
        lab_t = yield self.env.process(self._book_from_pool(
            self.mat_lab_slots, self.mat_exam_lock,
            not_before=self.env.now + self.cfg.mat_lab_lead_days * 24 * 60,
            waitslot_key='mat_lab_lead', pid=patient.pid))
        if us_t is None or lab_t is None:
            self._complete_post(patient)
            return
        exams_ready = max(us_t + self.cfg.mat_us_duration_min,
                          lab_t + self.cfg.mat_lab_duration_min)

        special_used = False
        while True:
            if (not special_used) and (self.rng.random() < self.cfg.preq_special_treat_p):
                special_used = True
                yield self.env.timeout(self.cfg.preq_special_treat_days * 24 * 60)
                continue
            done = self.env.event()
            need = PostNeed(
                nid=next(self._post_need_id), patient=patient,
                not_before=exams_ready, source='matrona', kind='preq',
                requires_labs=False, requires_us=False,
                empty_repeat_left=0, rebooks_left=self.cfg.max_control_rebooks,
                queue_priority=patient.priority, done_event=done,
            )
            self._enqueue_need(need)
            self._set_patient_state(patient.pid, 'waiting_post_need_preq')
            outcome = yield done
            if outcome != 'attended':
                self._complete_post(patient)
                return
            break

        if tr:
            tr['preq_completed_at'] = float(self.env.now)
            tr['follows_surgical_path'] = True
            tr['surgical_entry_source'] = 'preq'

        self._set_patient_state(patient.pid, 'transition_preq_to_quir')
        yield self.env.process(self.route_quir(patient, from_preq=True))

    def route_quir(self, patient, from_preq: bool):
        tr = self.patient_trace.get(patient.pid)
        if from_preq:
            if tr:
                tr['entered_quir_at'] = float(self.env.now)
                tr['surgical_entry_source'] = 'preq'
            self._set_patient_state(patient.pid, 'quir_from_preq')
        else:
            if tr:
                tr['post_route'] = 'quir'
                tr['entered_quir_at'] = float(self.env.now)
                tr['surgical_entry_source'] = 'quir_direct'
            yield self.env.process(self._matrona_work(
                self.cfg.matrona_review_min_low,
                self.cfg.matrona_review_min_high,
                pid=patient.pid))
            us_t = yield self.env.process(self._book_from_pool(
                self.mat_us_slots, self.mat_exam_lock,
                not_before=self.env.now + self.cfg.mat_us_lead_days * 24 * 60,
                waitslot_key='mat_us_lead', pid=patient.pid))
            lab_t = yield self.env.process(self._book_from_pool(
                self.mat_lab_slots, self.mat_exam_lock,
                not_before=self.env.now + self.cfg.mat_lab_lead_days * 24 * 60,
                waitslot_key='mat_lab_lead', pid=patient.pid))
            if us_t is None or lab_t is None:
                self._complete_post(patient)
                return
            exams_ready = max(us_t + self.cfg.mat_us_duration_min,
                              lab_t + self.cfg.mat_lab_duration_min)
            t_follow = exams_ready + 2 * 24 * 60
            if self.env.now < t_follow:
                yield self.env.timeout(t_follow - self.env.now)
            yield self.env.process(self._matrona_fixed(20.0, pid=patient.pid))

        altered = (self.rng.random() < self.cfg.quir_altered_p)
        if altered:
            self.kpis.quir_altered_total += 1
            if self.rng.random() < self.cfg.quir_altered_option_specialist_p:
                self.kpis.quir_altered_treatment += 1
                self._set_patient_state(patient.pid, 'quir_altered_treatment')
                yield self.env.timeout(self.cfg.quir_altered_treatment_days * 24 * 60)

                lab_t2 = yield self.env.process(self._book_from_pool(
                    self.mat_lab_slots, self.mat_exam_lock,
                    not_before=self.env.now + self.cfg.mat_lab_lead_days * 24 * 60,
                    waitslot_key='mat_lab_lead_quir_altered', pid=patient.pid))
                if lab_t2 is None:
                    self._complete_post(patient)
                    return
                ready2 = lab_t2 + self.cfg.mat_lab_duration_min
                if self.env.now < ready2:
                    yield self.env.timeout(ready2 - self.env.now)

                done = self.env.event()
                need = PostNeed(
                    nid=next(self._post_need_id), patient=patient,
                    not_before=self.env.now, source='matrona', kind='quir',
                    requires_labs=False, requires_us=False,
                    empty_repeat_left=0, rebooks_left=self.cfg.max_control_rebooks,
                    queue_priority=patient.priority, done_event=done,
                )
                self._enqueue_need(need)
                self._set_patient_state(patient.pid, 'waiting_post_need_quir')
                outcome = yield done
                if outcome != 'attended':
                    self._complete_post(patient)
                    return
            else:
                self.kpis.quir_altered_anesth += 1
                yield self.env.process(self._matrona_fixed(5.0, pid=patient.pid))
                lead_days = float(self.rng.uniform(
                    self.cfg.anesth_lead_days_min,
                    self.cfg.anesth_lead_days_max))
                t_an = self.env.now + lead_days * 24 * 60
                self._trace_add_waitslot(patient.pid, 'anesth_lead', t_an - self.env.now)
                if self.env.now < t_an:
                    yield self.env.timeout(t_an - self.env.now)
                yield self.env.process(self._service_loop(
                    self.anesthesist, self.cfg.anesth_visit_min, 'work',
                    pid=patient.pid, waitq_key='anesthesist'))
                self.kpis.post_anesth_attended += 1

            yield self.env.process(self._matrona_work(
                self.cfg.matrona_review_min_low,
                self.cfg.matrona_review_min_high,
                pid=patient.pid))
        else:
            yield self.env.process(self._matrona_fixed(20.0, pid=patient.pid))

        self._set_patient_state(patient.pid, 'waiting_surgery')
        t0_surg = self.env.now

        crs_lead = float(self.rng.uniform(
            self.cfg.crs_lead_days_min,
            self.cfg.crs_lead_days_max)) * 24 * 60
        self._trace_add_waitslot(patient.pid, 'crs_lead', crs_lead)
        yield self.env.timeout(crs_lead)

        self.surgery_waiting_open += 1
        yield self.surgery_tokens.get(1)
        self.surgery_waiting_open = max(0, self.surgery_waiting_open - 1)
        self._trace_add_waitslot(patient.pid, 'surgery_wait', self.env.now - t0_surg - crs_lead)
        self._set_patient_state(patient.pid, 'after_surgery')
        yield self.env.process(self._after_surgery(patient))

    def _after_surgery(self, patient):
        req_days = float(self.rng.uniform(
            self.cfg.becado_request_after_surgery_days_min,
            self.cfg.becado_request_after_surgery_days_max))
        t_req = self.env.now + req_days * 24 * 60
        if self.env.now < t_req:
            yield self.env.timeout(t_req - self.env.now)
        t_book = self.env.now + float(self.rng.uniform(
            0, self.cfg.becado_book_within_days)) * 24 * 60
        self._trace_add_waitslot(patient.pid, 'becado_lead', t_book - self.env.now)
        if self.env.now < t_book:
            yield self.env.timeout(t_book - self.env.now)
        yield self.env.process(self._service_loop(
            self.becado, self.cfg.becado_visit_min, 'work',
            pid=patient.pid, waitq_key='becado'))
        self.kpis.post_becado_attended += 1

        done = self.env.event()
        need = PostNeed(
            nid=next(self._post_need_id), patient=patient,
            not_before=self.env.now + 30 * 24 * 60, source='ugd', kind='postop',
            requires_labs=False, requires_us=False,
            empty_repeat_left=0, rebooks_left=self.cfg.max_control_rebooks,
            queue_priority=Priority.HIGH, done_event=done,
        )
        self._enqueue_need(need)
        self._set_patient_state(patient.pid, 'waiting_post_need_postop')
        outcome = yield done
        if outcome != 'attended':
            self._complete_post(patient)
            return
        yield self.env.process(self._post_discharge_or_monthly(patient))

    def route_surgery_waitlist(self, patient):
        self._set_patient_state(patient.pid, 'waiting_surgery')
        t0 = self.env.now

        crs_lead = float(self.rng.uniform(
            self.cfg.crs_lead_days_min,
            self.cfg.crs_lead_days_max)) * 24 * 60
        self._trace_add_waitslot(patient.pid, 'crs_lead', crs_lead)
        yield self.env.timeout(crs_lead)

        self.surgery_waiting_open += 1
        yield self.surgery_tokens.get(1)
        self.surgery_waiting_open = max(0, self.surgery_waiting_open - 1)
        self._trace_add_waitslot(patient.pid, 'surgery_wait', self.env.now - t0 - crs_lead)
        self._set_patient_state(patient.pid, 'after_surgery')
        yield self.env.process(self._after_surgery(patient))

    def _post_discharge_or_monthly(self, patient):
        if not patient.endometriosis:
            self._complete_post(patient)
            return
        if self.rng.random() < 0.10:
            self._complete_post(patient)
            return
        done = self.env.event()
        need = PostNeed(
            nid=next(self._post_need_id), patient=patient,
            not_before=self.env.now + 30 * 24 * 60, source='ugd', kind='monthly',
            requires_labs=False, requires_us=False,
            empty_repeat_left=0, rebooks_left=self.cfg.max_control_rebooks,
            queue_priority=patient.priority, done_event=done,
        )
        self._enqueue_need(need)
        self._set_patient_state(patient.pid, 'waiting_post_need_monthly')
        outcome = yield done
        if outcome != 'attended':
            self._complete_post(patient)
            return
        self._complete_post(patient)

    def _complete_post(self, patient):
        if patient.pid not in self.post_in_process:
            return
        self.post_in_process.remove(patient.pid)
        self.kpis.post_completed += 1
        self.kpis.post_completed_by_pri[patient.priority] += 1
        self.post_completed_cum += 1
        t0 = self.post_start_time.get(patient.pid)
        if t0 is not None:
            self.kpis.post_total_time_all.append(self.env.now - t0)
        self.kpis.total_time_full_all.append(self.env.now - patient.enqueued_at)
        tr = self.patient_trace.get(patient.pid)
        if tr:
            tr['post_done_at'] = float(self.env.now)
        self._set_patient_state(patient.pid, 'post_completed')

    def _enqueue_need(self, need: PostNeed):
        pr = need.queue_priority
        if need.source == 'matrona':
            if pr == Priority.HIGH:   self.post_wait_high_matrona.append(need)
            elif pr == Priority.MID:  self.post_wait_mid_matrona.append(need)
            else:                     self.post_wait_low_matrona.append(need)
        else:
            if pr == Priority.HIGH:   self.post_wait_high_ugd.append(need)
            elif pr == Priority.MID:  self.post_wait_mid_ugd.append(need)
            else:                     self.post_wait_low_ugd.append(need)

    def _pick_need(self, qH, qM, qL):
        now = self.env.now
        def eligible(q): return [n for n in q if n.not_before <= now]
        for q in (qH, qM, qL):
            cand = eligible(q)
            if cand:
                return max(cand, key=lambda n: (now - n.patient.enqueued_at))
        return None

    def _remove_need(self, need: PostNeed, qH, qM, qL):
        pr = need.queue_priority
        q = qH if pr == Priority.HIGH else qM if pr == Priority.MID else qL
        try:
            q.remove(need)
        except ValueError:
            pass

    def agent_dispatcher_post_ugd(self):
        while True:
            if self.env.now >= self.cfg.sim_time_min:
                break
            t0 = (next_agent_minute(self.env.now, AGENT_BOOKING_WEEKDAYS_POST, self.cfg)
                  if self.cfg.agent_limit_to_weekdays
                  else CAL_UGD.next_work_minute(self.env.now))
            if t0 > self.env.now:
                yield self.env.timeout(t0 - self.env.now)

            while self.post_control_slots and self.post_control_slots[0] <= self.env.now:
                self.post_control_slots.popleft()
            if not self.post_control_slots:
                yield self.env.timeout(10)
                continue
            need = self._pick_need(
                self.post_wait_high_ugd, self.post_wait_mid_ugd, self.post_wait_low_ugd)
            if need is None:
                yield self.env.timeout(10)
                continue

            t_q_lock = self.env.now
            lock = self.post_slot_lock.request(priority=1)
            yield lock
            self._trace_add_waitq(need.patient.pid, 'post_slot_lock_ugd', self.env.now - t_q_lock)
            try:
                t_ready = self._next_with_capacity(
                    self.env.now,
                    self.cfg.contact_attempt_min + self.cfg.case_time_short_min,
                    'agent', AGENT_BOOKING_WEEKDAYS_POST)
                if t_ready > self.env.now:
                    yield self.env.timeout(t_ready - self.env.now)
                t_q_agent = self.env.now
                with self.agent.request() as req:
                    yield req
                    if (self._calendar_end_today(self.env.now, 'agent', AGENT_BOOKING_WEEKDAYS_POST)
                            - self.env.now < (self.cfg.contact_attempt_min + self.cfg.case_time_short_min - 1e-9)):
                        continue
                    self._trace_add_waitq(need.patient.pid, 'agent_post', self.env.now - t_q_agent)
                    self._mark_active_non_surgery(need.patient.pid, True)
                    try:
                        yield self.env.timeout(self.cfg.contact_attempt_min)
                    finally:
                        self._mark_active_non_surgery(need.patient.pid, False)
                    self.kpis.post_agent_time_total += self.cfg.contact_attempt_min

                    if self.rng.random() < self.cfg.not_contactable_p:
                        self.kpis.not_contactable += 1
                        self.contact_attempts[need.patient.pid] += 1
                        if self.contact_attempts[need.patient.pid] >= self.cfg.max_contact_attempts:
                            self._remove_need(need, self.post_wait_high_ugd,
                                              self.post_wait_mid_ugd, self.post_wait_low_ugd)
                            self.kpis.unreachable_removed += 1
                            self.cum_removed_total += 1
                            self.cum_removed_by_pri[need.patient.priority] += 1
                            self._resolve_done_event(need, 'dropped')
                            self._complete_post(need.patient)
                        else:
                            self._set_patient_state(
                                need.patient.pid, f'waiting_post_need_{need.kind}')
                        continue

                    earliest_ok = max(self.env.now, need.not_before)
                    chosen_idx = None
                    for i, s in enumerate(self.post_control_slots):
                        if s >= earliest_ok:
                            chosen_idx = i
                            break
                    if chosen_idx is None:
                        yield self.env.timeout(10)
                        continue
                    slot_t     = self.post_control_slots[chosen_idx]
                    lead_hours = max(0.0, (slot_t - self.env.now) / 60.0)
                    is_long    = (lead_hours >= self.cfg.lead_short_hours)
                    call_time  = (self.cfg.case_time_short_min if not is_long
                                  else self.cfg.case_time_long_min)
                    self._mark_active_non_surgery(need.patient.pid, True)
                    try:
                        yield self.env.timeout(call_time)
                    finally:
                        self._mark_active_non_surgery(need.patient.pid, False)
                    self.kpis.post_agent_time_total += call_time

                while self.post_control_slots and self.post_control_slots[0] <= self.env.now:
                    self.post_control_slots.popleft()
                if chosen_idx >= len(self.post_control_slots):
                    continue
                slot_t = self.post_control_slots[chosen_idx]
                del self.post_control_slots[chosen_idx]
                self._remove_need(need, self.post_wait_high_ugd,
                                  self.post_wait_mid_ugd, self.post_wait_low_ugd)
                self.kpis.post_control_bookings += 1
                self.kpis.post_control_bookings_by_source['ugd'] += 1
                self._trace_add_waitslot(
                    need.patient.pid, 'post_control_lead', slot_t - self.env.now)
                self.post_booked_open[need.kind] += 1
                self._set_patient_state(need.patient.pid, f'booked_post_need_{need.kind}')
                self.env.process(self.post_control_appointment_event(need, slot_t))
            finally:
                try:
                    self.post_slot_lock.release(lock)
                except Exception:
                    pass

    def dispatcher_post_matrona(self):
        while True:
            if self.env.now >= self.cfg.sim_time_min:
                break
            t0 = CAL_MATRONA.next_work_minute(self.env.now)
            if t0 > self.env.now:
                yield self.env.timeout(t0 - self.env.now)
            while self.post_control_slots and self.post_control_slots[0] <= self.env.now:
                self.post_control_slots.popleft()
            if not self.post_control_slots:
                yield self.env.timeout(10)
                continue
            need = self._pick_need(self.post_wait_high_matrona,
                                   self.post_wait_mid_matrona,
                                   self.post_wait_low_matrona)
            if need is None:
                yield self.env.timeout(10)
                continue
            t_q_lock = self.env.now
            lock = self.post_slot_lock.request(priority=0)
            yield lock
            self._trace_add_waitq(
                need.patient.pid, 'post_slot_lock_matrona', self.env.now - t_q_lock)
            try:
                yield from self._matrona_fixed(self.cfg.matrona_booking_min,
                                               pid=need.patient.pid)
                earliest_ok = max(self.env.now, need.not_before)
                chosen_idx = None
                for i, s in enumerate(self.post_control_slots):
                    if s >= earliest_ok:
                        chosen_idx = i
                        break
                if chosen_idx is None:
                    yield self.env.timeout(10)
                    continue
                while self.post_control_slots and self.post_control_slots[0] <= self.env.now:
                    self.post_control_slots.popleft()
                if chosen_idx >= len(self.post_control_slots):
                    continue
                slot_t = self.post_control_slots[chosen_idx]
                del self.post_control_slots[chosen_idx]
                self._remove_need(need, self.post_wait_high_matrona,
                                  self.post_wait_mid_matrona, self.post_wait_low_matrona)
                self.kpis.post_control_bookings += 1
                self.kpis.post_control_bookings_by_source['matrona'] += 1
                self._trace_add_waitslot(
                    need.patient.pid, 'post_control_lead', slot_t - self.env.now)
                self.post_booked_open[need.kind] += 1
                self._set_patient_state(need.patient.pid, f'booked_post_need_{need.kind}')
                self.env.process(self.post_control_appointment_event(need, slot_t))
            finally:
                try:
                    self.post_slot_lock.release(lock)
                except Exception:
                    pass

    def post_control_appointment_event(self, need, slot_t):
        if self.env.now < slot_t:
            yield self.env.timeout(slot_t - self.env.now)
        self._set_patient_state(need.patient.pid, f'in_post_need_{need.kind}')
        self.post_booked_open[need.kind] = max(0, self.post_booked_open[need.kind] - 1)

        if need.kind == 'control' and self.rng.random() < self.cfg.blocked_pct_post_control:
            self.kpis.post_control_blocked += 1
            if need.rebooks_left > 0:
                new_need = PostNeed(
                    nid=next(self._post_need_id), patient=need.patient,
                    not_before=self.next_post_publish_after(self.env.now),
                    source=need.source, kind=need.kind,
                    requires_labs=need.requires_labs, requires_us=need.requires_us,
                    empty_repeat_left=need.empty_repeat_left,
                    rebooks_left=need.rebooks_left - 1,
                    queue_priority=need.queue_priority,
                    done_event=need.done_event,
                )
                self._enqueue_need(new_need)
                self._set_patient_state(need.patient.pid, f'waiting_post_need_{need.kind}')
            else:
                self._resolve_done_event(need, 'dropped')
            return

        if self.rng.random() < self.cfg.nsp_matrona_p:
            self.kpis.post_control_nsp += 1
            self._mark_active_non_surgery(need.patient.pid, True)
            try:
                yield self.env.timeout(self.cfg.consult_duration_min)
            finally:
                self._mark_active_non_surgery(need.patient.pid, False)
            if need.rebooks_left > 0:
                new_need = PostNeed(
                    nid=next(self._post_need_id), patient=need.patient,
                    not_before=self.next_post_publish_after(self.env.now),
                    source=need.source, kind=need.kind,
                    requires_labs=need.requires_labs, requires_us=need.requires_us,
                    empty_repeat_left=need.empty_repeat_left,
                    rebooks_left=need.rebooks_left - 1,
                    queue_priority=need.queue_priority,
                    done_event=need.done_event,
                )
                self._enqueue_need(new_need)
                self._set_patient_state(need.patient.pid, f'waiting_post_need_{need.kind}')
            else:
                self._resolve_done_event(need, 'dropped')
            return

        if (need.source == 'ugd' and need.kind == 'control'
                and need.requires_labs and need.empty_repeat_left > 0):
            if self.rng.random() < self.cfg.empty_control_p_ugd:
                self.kpis.post_control_empty += 1
                self._mark_active_non_surgery(need.patient.pid, True)
                try:
                    yield self.env.timeout(self.cfg.consult_duration_min)
                finally:
                    self._mark_active_non_surgery(need.patient.pid, False)
                us_t = yield self.env.process(self._book_from_pool(
                    self.ugd_us_slots, self.ugd_us_lock,
                    not_before=self.env.now, waitslot_key='ugd_us_lead',
                    pid=need.patient.pid))
                if us_t is None:
                    self._resolve_done_event(need, 'dropped')
                    return
                ready = us_t + self.cfg.ugd_us_duration_min
                if need.rebooks_left > 0:
                    new_need = PostNeed(
                        nid=next(self._post_need_id), patient=need.patient,
                        not_before=ready, source=need.source, kind=need.kind,
                        requires_labs=need.requires_labs, requires_us=need.requires_us,
                        empty_repeat_left=need.empty_repeat_left - 1,
                        rebooks_left=need.rebooks_left - 1,
                        queue_priority=need.queue_priority,
                        done_event=need.done_event,
                    )
                    self._enqueue_need(new_need)
                    self._set_patient_state(need.patient.pid, f'waiting_post_need_{need.kind}')
                else:
                    self._resolve_done_event(need, 'dropped')
                return

        day = int(self.env.now // (24 * 60))
        end_of_day = day * 24 * 60 + self.cfg.work_end_h * 60
        if end_of_day - self.env.now < self.cfg.consult_duration_min - 1e-9:
            if need.rebooks_left > 0:
                new_need = PostNeed(
                    nid=next(self._post_need_id), patient=need.patient,
                    not_before=self.next_post_publish_after(self.env.now),
                    source=need.source, kind=need.kind,
                    requires_labs=need.requires_labs, requires_us=need.requires_us,
                    empty_repeat_left=need.empty_repeat_left,
                    rebooks_left=need.rebooks_left - 1,
                    queue_priority=need.queue_priority,
                    done_event=need.done_event,
                )
                self._enqueue_need(new_need)
                self._set_patient_state(need.patient.pid, f'waiting_post_need_{need.kind}')
            else:
                self._resolve_done_event(need, 'dropped')
            return

        self._mark_active_non_surgery(need.patient.pid, True)
        try:
            yield self.env.timeout(self.cfg.consult_duration_min)
        finally:
            self._mark_active_non_surgery(need.patient.pid, False)
        self.kpis.post_control_attended += 1
        self.kpis.post_control_att_by_source[need.source] += 1
        self.kpis.post_control_minutes_used += self.cfg.consult_duration_min
        self._resolve_done_event(need, 'attended')

    # ── Loops de publicación de slots ───────────────────────────────────────
    def weekly_pre_publish_loop(self):
        week = 0
        while True:
            week_start = week * 7 * 24 * 60
            if week_start >= self.cfg.sim_time_min:
                break
            monday_8am = week_start + self.cfg.work_start_h * 60
            pub_t = max(0, CAL_UGD.workdays_before(monday_8am, self.cfg.publish_lead_workdays))
            if self.env.now < pub_t:
                yield self.env.timeout(pub_t - self.env.now)

            slots_expired = len(self.slot_times)
            if not hasattr(self.kpis, 'slots_expired_total'):
                self.kpis.slots_expired_total = 0
            self.kpis.slots_expired_total += slots_expired

            self.slot_times.clear()
            self.published_at = self.env.now
            self.earliest_bookable_time = self.published_at + self.cfg.min_lead_publish_hours * 60

            if self.cfg.use_fixed_weekly_capacity:
                week_capacity_raw = int(self.cfg.fixed_weekly_capacity)
            else:
                week_capacity_raw = int(max(0, round(
                    0.5 + self.rng.gamma(self.cfg.gamma_shape, self.cfg.gamma_scale))))

            day_caps = {d: int(round(week_capacity_raw * self.cfg.day_share.get(d, 0.0)))
                        for d in range(5)}
            diff = week_capacity_raw - sum(day_caps.values())
            order = sorted(day_caps, key=lambda k: self.cfg.day_share.get(k, 0.0), reverse=True)
            i = 0
            while diff != 0 and order:
                k = order[i % len(order)]
                if diff > 0:
                    day_caps[k] += 1; diff -= 1
                elif day_caps[k] > 0:
                    day_caps[k] -= 1; diff += 1
                i += 1

            for d in range(5):
                cap_d = day_caps[d]
                if cap_d <= 0:
                    continue
                day_begin = week_start + d * 24 * 60
                start = day_begin + self.cfg.work_start_h * 60
                end   = day_begin + self.cfg.work_end_h   * 60
                max_slots = int((end - start) // self.cfg.consult_duration_min)
                n = min(cap_d, max_slots)
                if n <= 0:
                    continue
                step = (end - start - self.cfg.consult_duration_min) / (n - 1) if n > 1 else 0
                for j in range(n):
                    s = int(start + j * step)
                    if s + self.cfg.consult_duration_min <= end and s < self.cfg.sim_time_min:
                        self.slot_times.append(s)

            self.kpis.slot_minutes_published += len(self.slot_times) * self.cfg.consult_duration_min
            self.week_caps_used[week] = {"long": 0, "short": 0}
            week += 1

    def weekly_post_publish_loop(self):
        week = 0
        while True:
            week_start = week * 7 * 24 * 60
            if week_start >= self.cfg.sim_time_min:
                break
            monday_8am = week_start + self.cfg.work_start_h * 60
            pub_t = max(0, CAL_UGD.workdays_before(monday_8am, self.cfg.publish_lead_workdays))
            if self.env.now < pub_t:
                yield self.env.timeout(pub_t - self.env.now)

            post_slots_expired = len(self.post_control_slots)
            if not hasattr(self.kpis, 'post_slots_expired_total'):
                self.kpis.post_slots_expired_total = 0
            self.kpis.post_slots_expired_total += post_slots_expired

            self.post_control_slots.clear()

            if self.cfg.use_fixed_post_control_hours:
                week_capacity_raw = int(self.cfg.fixed_post_control_capacity)
            else:
                hours = (self.cfg.post_control_hours_base
                        + self.cfg.post_control_hours_mult
                        * self.rng.beta(self.cfg.post_control_beta_a,
                                        self.cfg.post_control_beta_b))
                week_capacity_raw = int(max(0.0, hours * 60.0) // self.cfg.consult_duration_min)

            day_caps = {d: int(round(week_capacity_raw * self.cfg.day_share_post.get(d, 0.0)))
                        for d in range(5)}
            diff = week_capacity_raw - sum(day_caps.values())
            order = sorted(day_caps, key=lambda k: self.cfg.day_share_post.get(k, 0.0), reverse=True)
            i = 0
            while diff != 0 and order:
                k = order[i % len(order)]
                if diff > 0:
                    day_caps[k] += 1; diff -= 1
                elif day_caps[k] > 0:
                    day_caps[k] -= 1; diff += 1
                i += 1

            for d in range(5):
                cap_d = day_caps[d]
                if cap_d <= 0:
                    continue
                day_begin = week_start + d * 24 * 60
                start = day_begin + self.cfg.work_start_h * 60
                end   = day_begin + self.cfg.work_end_h   * 60
                max_slots = int((end - start) // self.cfg.consult_duration_min)
                n = min(cap_d, max_slots)
                if n <= 0:
                    continue
                step = (end - start - self.cfg.consult_duration_min) / (n - 1) if n > 1 else 0
                for j in range(n):
                    s = int(start + j * step)
                    if s + self.cfg.consult_duration_min <= end and s < self.cfg.sim_time_min:
                        self.post_control_slots.append(s)

            self.kpis.post_control_minutes_published += (
                len(self.post_control_slots) * self.cfg.consult_duration_min)
            week += 1

    def weekly_ugd_us_slots_loop(self):
        week = 0
        while True:
            week_start = week * 7 * 24 * 60
            if week_start >= self.cfg.sim_time_min:
                break
            t_gen = week_start + self.cfg.work_start_h * 60
            if self.env.now < t_gen:
                yield self.env.timeout(t_gen - self.env.now)
            self.ugd_us_slots.clear()
            ugd_us_eff = int(round(self.cfg.ugd_us_per_week * (1.0 + self.cfg.ugd_us_extra_pct)))
            per_day = [ugd_us_eff // 5] * 5
            for i in range(ugd_us_eff % 5):
                per_day[i] += 1
            for d in range(5):
                n = per_day[d]
                if n <= 0:
                    continue
                day_begin = week_start + d * 24 * 60
                start = day_begin + self.cfg.work_start_h * 60
                end   = day_begin + self.cfg.work_end_h   * 60
                max_slots = int((end - start) // self.cfg.ugd_us_duration_min)
                n = min(n, max_slots)
                if n <= 0:
                    continue
                step = (end - start - self.cfg.ugd_us_duration_min) / (n - 1) if n > 1 else 0
                for j in range(n):
                    s = int(start + j * step)
                    if s + self.cfg.ugd_us_duration_min <= end and s < self.cfg.sim_time_min:
                        self.ugd_us_slots.append(s)
            week += 1

    def weekly_ugd_lab_slots_loop(self):
        week = 0
        while True:
            week_start = week * 7 * 24 * 60
            if week_start >= self.cfg.sim_time_min:
                break
            t_gen = week_start + self.cfg.work_start_h * 60
            if self.env.now < t_gen:
                yield self.env.timeout(t_gen - self.env.now)
            self.ugd_lab_slots.clear()
            per_day = [self.cfg.ugd_lab_per_week // 5] * 5
            for i in range(self.cfg.ugd_lab_per_week % 5):
                per_day[i] += 1
            for d in range(5):
                n = per_day[d]
                if n <= 0:
                    continue
                day_begin = week_start + d * 24 * 60
                start = day_begin + self.cfg.work_start_h * 60
                end   = day_begin + self.cfg.work_end_h   * 60
                max_slots = int((end - start) // self.cfg.ugd_lab_duration_min)
                n = min(n, max_slots)
                if n <= 0:
                    continue
                step = (end - start - self.cfg.ugd_lab_duration_min) / (n - 1) if n > 1 else 0
                for j in range(n):
                    s = int(start + j * step)
                    if s + self.cfg.ugd_lab_duration_min <= end and s < self.cfg.sim_time_min:
                        self.ugd_lab_slots.append(s)
            week += 1

    def weekly_mat_exam_slots_loop(self):
        week = 0
        while True:
            week_start = week * 7 * 24 * 60
            if week_start >= self.cfg.sim_time_min:
                break
            t_gen = week_start + self.cfg.work_start_h * 60
            if self.env.now < t_gen:
                yield self.env.timeout(t_gen - self.env.now)
            self.mat_us_slots.clear()
            self.mat_lab_slots.clear()

            per_day_us = [self.cfg.mat_us_per_week // 5] * 5
            for i in range(self.cfg.mat_us_per_week % 5):
                per_day_us[i] += 1
            for d in range(5):
                n = per_day_us[d]
                if n <= 0:
                    continue
                day_begin = week_start + d * 24 * 60
                start = day_begin + self.cfg.work_start_h * 60
                end   = day_begin + self.cfg.work_end_h   * 60
                max_slots = int((end - start) // self.cfg.mat_us_duration_min)
                n = min(n, max_slots)
                if n <= 0:
                    continue
                step = (end - start - self.cfg.mat_us_duration_min) / (n - 1) if n > 1 else 0
                for j in range(n):
                    s = int(start + j * step)
                    if s + self.cfg.mat_us_duration_min <= end and s < self.cfg.sim_time_min:
                        self.mat_us_slots.append(s)

            per_day_lab = [self.cfg.mat_lab_per_week // 5] * 5
            for i in range(self.cfg.mat_lab_per_week % 5):
                per_day_lab[i] += 1
            for d in range(5):
                n = per_day_lab[d]
                if n <= 0:
                    continue
                day_begin = week_start + d * 24 * 60
                start = day_begin + self.cfg.work_start_h * 60
                end   = day_begin + self.cfg.work_end_h   * 60
                max_slots = int((end - start) // self.cfg.mat_lab_duration_min)
                n = min(n, max_slots)
                if n <= 0:
                    continue
                step = (end - start - self.cfg.mat_lab_duration_min) / (n - 1) if n > 1 else 0
                for j in range(n):
                    s = int(start + j * step)
                    if s + self.cfg.mat_lab_duration_min <= end and s < self.cfg.sim_time_min:
                        self.mat_lab_slots.append(s)
            week += 1

    def weekly_surgery_tokens_loop(self):
        week = 0
        while True:
            week_start = week * 7 * 24 * 60
            if week_start >= self.cfg.sim_time_min:
                break
            total = int(max(0, round(self.cfg.surg_throughput_per_week)))
            day_caps = {d: int(round(total / 5.0)) for d in range(5)}
            diff = total - sum(day_caps.values())
            d = 0
            while diff != 0:
                k = d % 5
                if diff > 0:
                    day_caps[k] += 1; diff -= 1
                elif day_caps[k] > 0:
                    day_caps[k] -= 1; diff += 1
                d += 1
            for d in range(5):
                t_put = week_start + d * 24 * 60 + self.cfg.work_start_h * 60
                if t_put >= self.cfg.sim_time_min:
                    continue
                if self.env.now < t_put:
                    yield self.env.timeout(t_put - self.env.now)
                n = int(day_caps[d])
                if n > 0:
                    yield self.surgery_tokens.put(n)
            week += 1

    def preq_weekly_selector_loop(self):
        week = 0
        while True:
            week_start = week * 7 * 24 * 60
            if week_start >= self.cfg.sim_time_min:
                break
            t_sel = week_start + self.cfg.work_start_h * 60
            if self.env.now < t_sel:
                yield self.env.timeout(t_sel - self.env.now)
            for _ in range(self.cfg.preq_select_per_week):
                if not self.preq_queue:
                    break
                now = self.env.now
                def rank(p):
                    pr = 0 if p.priority == Priority.HIGH else 1 if p.priority == Priority.MID else 2
                    waited = now - self.preq_enq_time.get(p.pid, now)
                    return (pr, -waited)
                p = min(self.preq_queue, key=rank)
                self.preq_queue.remove(p)
                ev = self.preq_select_events.pop(p.pid, None)
                if ev and not ev.triggered:
                    ev.succeed()
            week += 1

    # ── Monitor y series ────────────────────────────────────────────────────
    def daily_monitor_end_of_day(self):
        def next_close(t):
            day = int(t // (24 * 60))
            while True:
                if day % 7 < 5:
                    close = day * 24 * 60 + self.cfg.work_end_h * 60
                    if close >= t:
                        return close
                day += 1

        t = next_close(self.env.now)
        if t > self.env.now:
            yield self.env.timeout(t - self.env.now)
        while self.env.now < self.cfg.sim_time_min:
            self.push_timeseries_point()
            self.push_post_timeseries_point()
            self._daily_transition_validator()
            nxt = next_close(self.env.now + 1)
            if nxt >= self.cfg.sim_time_min:
                break
            yield self.env.timeout(nxt - self.env.now)

    def _daily_transition_validator(self):
        problems = {}
        for name, res in [('reconv', self.reconv), ('agent', self.agent),
                           ('matrona', self.matrona), ('anesthesist', self.anesthesist),
                           ('becado', self.becado)]:
            if len(res.users) > 0 and len(res.queue) > 0:
                problems[f'{name}_queue_backlog'] = len(res.queue)
        if len(self.active_non_surgery) > 3:
            problems['active_non_surgery_high'] = sorted(self.active_non_surgery)
        if problems:
            self.daily_transition_errors.append(
                {'time': float(self.env.now), 'problems': problems})

    def push_timeseries_point(self):
        self.kpis.ts_time_min.append(self.env.now)
        self.kpis.ts_cum_wait_total.append(self.cum_wait_total)
        self.kpis.ts_cum_att_total.append(self.att_total_cum)
        self.kpis.ts_diff_total.append(
            self.cum_wait_total - self.att_total_cum - self.cum_removed_total)

    def push_post_timeseries_point(self):
        self.kpis.ts_post_time_min.append(self.env.now)
        self.kpis.ts_post_entered_cum.append(self.post_entered_cum)
        self.kpis.ts_post_completed_cum.append(self.post_completed_cum)
        self.kpis.ts_post_in_process.append(len(self.post_in_process))
        self.kpis.ts_post_control_attended_cum.append(self.kpis.post_control_attended)

    def next_post_publish_after(self, t):
        w = week_index_from_time(t) + 1
        monday_8am = w * 7 * 24 * 60 + self.cfg.work_start_h * 60
        pub_t = CAL_UGD.workdays_before(monday_8am, self.cfg.publish_lead_workdays)
        return CAL_UGD.next_work_minute(max(pub_t, t))

    def first_waiting_now(self):
        return max(0, self.cum_wait_total - self.att_total_cum - self.cum_removed_total)

    def first_queue_only_now(self):
        return len(self.wait_high) + len(self.wait_mid) + len(self.wait_low)

    def _count_needs_kind(self, kind: str) -> int:
        def cnt(q): return sum(1 for n in q if getattr(n, 'kind', None) == kind)
        return (cnt(self.post_wait_high_matrona) + cnt(self.post_wait_mid_matrona) +
                cnt(self.post_wait_low_matrona)  + cnt(self.post_wait_high_ugd) +
                cnt(self.post_wait_mid_ugd)      + cnt(self.post_wait_low_ugd))

    def post_kind_waiting_now(self, kind: str) -> int:
        return int(self._count_needs_kind(kind) + int(self.post_booked_open.get(kind, 0)))

    def _count_control_route_active(self):
        monthly_states = {'waiting_post_need_monthly', 'booked_post_need_monthly',
                          'in_post_need_monthly'}
        return int(sum(1 for pid in self.post_in_process
                       if (self.patient_trace.get(pid, {}).get('post_route') == 'control'
                           and self.patient_state.get(pid, '') not in monthly_states)))

    def _count_quir_route_active(self):
        return int(sum(1 for pid in self.post_in_process
                       if (self.patient_trace.get(pid, {}).get('post_route') == 'quir'
                           or self.patient_trace.get(pid, {}).get('entered_quir_at') is not None)))

    def post_control_waiting_now(self):
        return self._count_control_route_active()

    def preq_waiting_now(self) -> int:
        return int(len(self.preq_queue) + self.post_kind_waiting_now("preq"))

    def quir_waiting_now(self) -> int:
        return self._count_quir_route_active()

    def total_in_system_now(self):
        return int(self.first_waiting_now() + len(self.post_in_process))

    def push_week_snapshot(self, week_idx: int):
        self.kpis.ts_week_idx.append(int(week_idx))
        self.kpis.ts_wl_first.append(int(self.first_waiting_now()))
        self.kpis.ts_wl_control.append(int(self.post_control_waiting_now()))
        self.kpis.ts_wl_preq.append(int(self.preq_waiting_now()))
        self.kpis.ts_wl_quir.append(int(self.quir_waiting_now()))

    def push_month_snapshot(self, month_idx: int):
        self.kpis.ts_month_idx.append(int(month_idx))
        self.kpis.ts_m_wl_first.append(int(self.first_waiting_now()))
        self.kpis.ts_m_wl_control.append(int(self.post_control_waiting_now()))
        self.kpis.ts_m_wl_preq.append(int(self.preq_waiting_now()))
        self.kpis.ts_m_wl_quir.append(int(self.quir_waiting_now()))

    def record_queue_snapshot_t0(self):
        yield self.env.timeout(0)
        self.push_week_snapshot(0)
        self.push_month_snapshot(0)

    def weekly_queue_monitor(self):
        month_idx = 0
        for w in range(self.cfg.weeks_to_simulate):
            day_friday  = w * 7 + 4
            t_fri_close = day_friday * 24 * 60 + self.cfg.work_end_h * 60
            if t_fri_close >= self.cfg.sim_time_min:
                break
            if self.env.now < t_fri_close:
                yield self.env.timeout(t_fri_close - self.env.now)
            week_idx = w + 1
            self.push_week_snapshot(week_idx)
            if week_idx % 4 == 0:
                month_idx += 1
                self.push_month_snapshot(month_idx)

    def finalize_validations(self):
        first_balance_ok = (self.cum_wait_total ==
                            self.att_total_cum + self.cum_removed_total + self.first_waiting_now())
        post_balance_ok  = (self.post_entered_cum ==
                            self.post_completed_cum + len(self.post_in_process))
        no_carryover_ok  = (len(self.daily_transition_errors) == 0)
        self.validation_checks = {
            'first_balance_ok':        bool(first_balance_ok),
            'post_balance_ok':         bool(post_balance_ok),
            'no_daily_carryover_ok':   bool(no_carryover_ok),
        }
        errors = []
        if not first_balance_ok:
            errors.append(
                f"Balance 1ra inconsistente: cum_wait={self.cum_wait_total}, "
                f"att={self.att_total_cum}, removed={self.cum_removed_total}, "
                f"open={self.first_waiting_now()}")
        if not post_balance_ok:
            errors.append(
                f"Balance post inconsistente: entered={self.post_entered_cum}, "
                f"completed={self.post_completed_cum}, in_process={len(self.post_in_process)}")
        if not no_carryover_ok:
            errors.append(
                f"{len(self.daily_transition_errors)} cierres con colas de recursos acumuladas (posible deadlock)")
        if errors:
            for e in errors:
                log.warning("Validación: %s", e)
        return {
            'checks': self.validation_checks,
            'errors': errors,
            'daily_transition_errors': self.daily_transition_errors,
        }


# ========================
# Una réplica
# ========================
def run_once(seed_offset=0, cfg: SimConfig = None):
    cfg = cfg or CFG
    rng = np.random.RandomState(cfg.random_seed_base + seed_offset)
    random.seed(cfg.random_seed_base + seed_offset)
    env   = simpy.Environment()
    model = ClinicModelAdjusted(env, rng, cfg)
    env.run(until=cfg.sim_time_min)
    validation_info = model.finalize_validations()
    k  = model.kpis
    pt = model.patient_trace

    first_done = [tr for tr in pt.values() if tr.get("first_done_at") is not None]
    post_done  = [tr for tr in pt.values() if tr.get("post_done_at")  is not None]

    FIRST_WAITQ_KEYS    = {"reconv", "agent_first"}
    FIRST_WAITSLOT_KEYS = {"first_appointment_lead"}
    POST_WAITQ_KEYS     = {"agent_post", "matrona", "anesthesist", "becado",
                           "pool_lock", "post_slot_lock_matrona", "post_slot_lock_ugd"}
    POST_WAITSLOT_KEYS  = {"post_control_lead", "mat_us_lead", "mat_lab_lead",
                           "ugd_us_lead", "anesth_lead", "becado_lead",
                           "surgery_wait", "mat_lab_lead_quir_altered", "crs_lead"}

    def sum_keys(dct, keys):
        return float(sum(dct.get(k, 0.0) for k in keys))

    pat_first_waitq    = [sum_keys(tr["waitq"],    FIRST_WAITQ_KEYS)    for tr in first_done]
    pat_first_waitslot = [sum_keys(tr["waitslot"], FIRST_WAITSLOT_KEYS) for tr in first_done]
    pat_post_waitq     = [sum_keys(tr["waitq"],    POST_WAITQ_KEYS)     for tr in post_done]
    pat_post_waitslot  = [sum_keys(tr["waitslot"], POST_WAITSLOT_KEYS)  for tr in post_done]
    pat_total_waitq    = [float(sum(tr["waitq"].values()))    for tr in post_done]
    pat_total_waitslot = [float(sum(tr["waitslot"].values())) for tr in post_done]

    def mins_to_days(x): return x / (60 * 24.0)
    def mean(lst): return float(sum(lst) / len(lst)) if lst else 0.0
    def med(lst):  return float(stats.median(lst))   if lst else 0.0

    res = {}
    res["first_bookings"]     = k.bookings
    res["first_attended"]     = k.attended
    res["first_cnu_count"]    = k.cnu_count
    res["post_entered"]       = k.post_entered
    res["post_completed"]     = k.post_completed
    res["post_ctrl_bookings"] = k.post_control_bookings
    res["post_ctrl_attended"] = k.post_control_attended
    res["post_ctrl_blocked"]  = k.post_control_blocked
    res["total_atenciones_first"] = k.attended
    res["total_atenciones_post"]  = k.post_completed
    res["total_atenciones"]       = k.attended + k.post_completed

    res["post_route_control"] = k.post_route_counts.get("control", 0)
    res["post_route_preq"]    = k.post_route_counts.get("preq", 0)
    res["post_route_quir"]    = k.post_route_counts.get("quir", 0)

    res["quir_altered_total"]      = k.quir_altered_total
    res["quir_altered_treatment"]  = k.quir_altered_treatment
    res["quir_altered_anesth"]     = k.quir_altered_anesth

    res["first_closed_count"]           = len(k.first_closed_total_time_all)
    res["first_attended_tts_count"]     = len(k.first_attended_total_time_all)
    res["tts_first_days_mean"]          = mins_to_days(mean(k.first_closed_total_time_all))
    res["tts_first_days_med"]           = mins_to_days(med(k.first_closed_total_time_all))
    res["tts_first_closed_days_mean"]   = mins_to_days(mean(k.first_closed_total_time_all))
    res["tts_first_closed_days_med"]    = mins_to_days(med(k.first_closed_total_time_all))
    res["tts_first_attended_days_mean"] = mins_to_days(mean(k.first_attended_total_time_all))
    res["tts_first_attended_days_med"]  = mins_to_days(med(k.first_attended_total_time_all))
    res["tts_first_closed_backlog_days_mean"]   = mins_to_days(mean(k.first_closed_backlog_time_all))
    res["tts_first_closed_process_days_mean"]   = mins_to_days(mean(k.first_closed_process_time_all))
    res["tts_first_attended_backlog_days_mean"] = mins_to_days(mean(k.first_attended_backlog_time_all))
    res["tts_first_attended_process_days_mean"] = mins_to_days(mean(k.first_attended_process_time_all))
    res["tts_first_backlog_days_mean"]  = res["tts_first_attended_backlog_days_mean"]
    res["tts_first_process_days_mean"]  = res["tts_first_attended_process_days_mean"]
    res["tts_post_days_mean"]  = mins_to_days(mean(k.post_total_time_all))
    res["tts_full_days_mean"]  = mins_to_days(mean(k.total_time_full_all))

    agent_work_total = (total_agent_minutes_until(cfg.sim_time_min, AGENT_BOOKING_WEEKDAYS_FIRST, cfg)
                        if cfg.agent_limit_to_weekdays
                        else CAL_UGD.total_work_minutes_until(cfg.sim_time_min))
    mat_work_total   = CAL_MATRONA.total_work_minutes_until(cfg.sim_time_min)
    res["agent_util_first_pct"]    = (k.agent_time_total / agent_work_total * 100) if agent_work_total else 0.0
    res["agent_util_post_pct"]     = (k.post_agent_time_total / agent_work_total * 100) if agent_work_total else 0.0
    res["agent_util_post_ugd_pct"] = res["agent_util_post_pct"]
    res["matrona_util_post_pct"]   = (k.post_matrona_time_total / mat_work_total * 100) if mat_work_total else 0.0
    res["matrona_util_pct"]        = res["matrona_util_post_pct"]
    res["spec_util_first_pct"]     = (k.specialist_minutes_used / k.slot_minutes_published * 100) if k.slot_minutes_published else 0.0
    res["spec_util_post_pct"]      = (k.post_control_minutes_used / k.post_control_minutes_published * 100) if k.post_control_minutes_published else 0.0
    res["slots_expired_first"]     = int(getattr(k, 'slots_expired_total', 0))
    res["slots_expired_post"]      = int(getattr(k, 'post_slots_expired_total', 0))

    for key, vals in k.waitq_minutes.items():
        res[f"waitq_{key}_min_mean"] = mean(vals)
    for key, vals in k.waitslot_minutes.items():
        res[f"waitslot_{key}_min_mean"] = mean(vals)

    res["pat_first_waitq_min_mean"]    = mean(pat_first_waitq)
    res["pat_first_waitslot_min_mean"] = mean(pat_first_waitslot)
    res["pat_post_waitq_min_mean"]     = mean(pat_post_waitq)
    res["pat_post_waitslot_min_mean"]  = mean(pat_post_waitslot)
    res["pat_total_waitq_min_mean"]    = mean(pat_total_waitq)
    res["pat_total_waitslot_min_mean"] = mean(pat_total_waitslot)

    res["_pat_total_wait_min"]     = pat_total_waitq
    res["_pat_total_waitslot_min"] = pat_total_waitslot

    res["_tts_first_total_min"]   = list(k.first_attended_total_time_all)
    res["_tts_first_backlog_min"] = list(k.first_attended_backlog_time_all)
    res["_tts_first_proceso_min"] = list(k.first_attended_process_time_all)

    nuevos_min = [
        k.first_attended_process_time_all[i]
        for i in range(len(k.first_attended_backlog_time_all))
        if k.first_attended_backlog_time_all[i] == 0.0
    ]
    res["_tts_first_nuevos_min"] = nuevos_min

    backlog_solo_min = [
        k.first_attended_total_time_all[i]
        for i in range(len(k.first_attended_backlog_time_all))
        if k.first_attended_backlog_time_all[i] > 0.0
    ]
    res["_tts_first_backlog_solo_min"] = backlog_solo_min

    sin_backlog_min = []
    for tr in pt.values():
        if tr.get("first_attended_at") is not None:
            t_enq = float(tr["t_enqueue"])
            t_att = float(tr["first_attended_at"])
            if t_enq > 0:
                mins = t_att - t_enq
                if mins >= 0:
                    sin_backlog_min.append(mins)
    res["_tts_first_sin_backlog_min"] = sin_backlog_min

    res["wl_first_end"]          = model.first_waiting_now()
    res["wl_control_end"]        = model.post_control_waiting_now()
    res["wl_total_end"]          = model.total_in_system_now()
    res["wl_first_queue_only_end"] = model.first_queue_only_now()

    res["validation_first_balance_ok"]      = 1 if validation_info["checks"]["first_balance_ok"]      else 0
    res["validation_post_balance_ok"]       = 1 if validation_info["checks"]["post_balance_ok"]       else 0
    res["validation_no_daily_carryover_ok"] = 1 if validation_info["checks"]["no_daily_carryover_ok"] else 0
    res["validation_error_count"]           = len(validation_info["errors"])
    res["_validation_info"] = validation_info

    res["_ts_wl_week"] = {
        "week": k.ts_week_idx, "wl_first": k.ts_wl_first,
        "wl_control": k.ts_wl_control, "wl_preq": k.ts_wl_preq, "wl_quir": k.ts_wl_quir,
    }
    res["_ts_wl_month"] = {
        "month": k.ts_month_idx, "wl_first": k.ts_m_wl_first,
        "wl_control": k.ts_m_wl_control, "wl_preq": k.ts_m_wl_preq, "wl_quir": k.ts_m_wl_quir,
    }

    post_waitslot_total_by_route = {"control": [], "preq": [], "quir": [], "other": []}
    post_waitslot_comp_by_route  = {r: defaultdict(list) for r in post_waitslot_total_by_route}
    for tr in post_done:
        r = tr.get("post_route") or "other"
        if r not in post_waitslot_total_by_route:
            r = "other"
        total_ws = float(sum(tr["waitslot"].values()))
        post_waitslot_total_by_route[r].append(total_ws)
        for k2, v2 in tr["waitslot"].items():
            post_waitslot_comp_by_route[r][k2].append(float(v2))

    res["_post_waitslot_total_by_route_min"] = post_waitslot_total_by_route
    post_waitslot_comp_mean_by_route = {}
    for r, dct in post_waitslot_comp_by_route.items():
        post_waitslot_comp_mean_by_route[r] = {
            k2: (float(np.mean(vs)) if vs else 0.0) for k2, vs in dct.items()}
    res["_post_waitslot_comp_mean_by_route_min"] = post_waitslot_comp_mean_by_route

    res["_ts_first"] = {
        "t": k.ts_time_min, "cum_wait": k.ts_cum_wait_total,
        "cum_att": k.ts_cum_att_total, "diff": k.ts_diff_total,
    }
    res["_ts_post"] = {
        "t": k.ts_post_time_min, "entered": k.ts_post_entered_cum,
        "completed": k.ts_post_completed_cum, "in_process": k.ts_post_in_process,
        "post_ctrl_att_cum": k.ts_post_control_attended_cum,
    }
    return res


# ========================
# Worker de nivel módulo (REQUERIDO para multiprocessing en Windows y macOS)
# Recibe (seed_offset, cfg_dict) → reconstruye SimConfig → llama run_once
# ========================
def _run_worker(args):
    """
    Wrapper pickleable para ProcessPoolExecutor.
    args = (seed_offset: int, cfg_dict: dict)
    """
    seed_offset, cfg_dict = args
    cfg = SimConfig(**cfg_dict)
    return run_once(seed_offset=seed_offset, cfg=cfg)


def _cfg_to_dict(cfg: SimConfig) -> dict:
    """Serializa SimConfig a dict plano (sin propiedades calculadas)."""
    return dataclasses.asdict(cfg)


# ========================
# Estadísticas
# ========================
def t_critical_975(df: int) -> float:
    if df <= 0:
        return 0.0
    try:
        from scipy.stats import t as student_t
        return float(student_t.ppf(0.975, df))
    except Exception:
        table = {
            1:12.706,2:4.303,3:3.182,4:2.776,5:2.571,6:2.447,7:2.365,8:2.306,
            9:2.262,10:2.228,11:2.201,12:2.179,13:2.160,14:2.145,15:2.131,
            16:2.120,17:2.110,18:2.101,19:2.093,20:2.086,21:2.080,22:2.074,
            23:2.069,24:2.064,25:2.060,26:2.056,27:2.052,28:2.048,29:2.045,
            30:2.042,40:2.021,60:2.000,120:1.980,
        }
        if df in table: return float(table[df])
        if df > 120: return 1.96
        lower = max(k for k in table if k < df)
        upper = min(k for k in table if k > df)
        frac  = (df - lower) / (upper - lower)
        return float(table[lower] + frac * (table[upper] - table[lower]))


def summarize(rep_results):
    keys = [k for k in rep_results[0].keys() if not k.startswith("_")]
    n    = len(rep_results)
    t95  = t_critical_975(n - 1) if n > 1 else 0.0
    agg  = {}
    for key in keys:
        vals = [r[key] for r in rep_results]
        mu   = float(np.mean(vals))
        sd   = float(np.std(vals, ddof=1)) if n > 1 else 0.0
        se   = float(sd / np.sqrt(n)) if n > 0 else 0.0
        err  = float(t95 * se) if n > 1 else 0.0
        agg[key] = {
            "mean": mu, "sd": sd, "se": se,
            "sample_error95_z": float(1.96 * se) if n > 1 else 0.0,
            "t_crit_95": float(t95),
            "sample_error95_t": err,
            "ci95_t_low":  float(mu - err),
            "ci95_t_high": float(mu + err),
            "n": int(n),
        }
    return agg


def export_summary_csv(agg: dict, filepath: str):
    import csv
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['kpi','mean','sd','se','sample_error95_z','t_crit_95',
                    'sample_error95_t','ci95_t_low','ci95_t_high','n'])
        for k in sorted(agg):
            r = agg[k]
            w.writerow([k, r['mean'], r['sd'], r['se'], r['sample_error95_z'],
                        r['t_crit_95'], r['sample_error95_t'],
                        r['ci95_t_low'], r['ci95_t_high'], int(r['n'])])


def export_replications_csv(rep_results, filepath: str):
    import csv
    if not rep_results:
        raise ValueError('rep_results vacío')
    keys = [k for k in rep_results[0] if not k.startswith('_')]
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(keys)
        for r in rep_results:
            w.writerow([r.get(k) for k in keys])


def export_summary_xlsx(agg: dict, filepath: str):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise RuntimeError('Instala openpyxl: pip install openpyxl') from e
    wb = Workbook()
    ws = wb.active
    ws.title = 'summary'
    ws.append(['kpi','mean','sd','se','sample_error95_z','t_crit_95',
               'sample_error95_t','ci95_t_low','ci95_t_high','n'])
    for k in sorted(agg):
        r = agg[k]
        ws.append([k, float(r['mean']), float(r['sd']), float(r['se']),
                   float(r['sample_error95_z']), float(r['t_crit_95']),
                   float(r['sample_error95_t']), float(r['ci95_t_low']),
                   float(r['ci95_t_high']), int(r['n'])])
    for col in range(1, 11):
        col_letter = get_column_letter(col)
        max_len = max((len(str(cell.value or '')) for cell in ws[col_letter]), default=10)
        ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))
    wb.save(filepath)


def pretty_print(agg, cfg=None):
    cfg = cfg or CFG
    def fmt(s, d=2):
        return (f"{s['mean']:.{d}f} ± {s['sd']:.{d}f} "
                f"| Err95-t {s['sample_error95_t']:.{d}f} | n={s['n']}")

    print(f"\n=== Resultados v3-paralelo ({cfg.weeks_to_simulate} semanas, "
          f"{cfg.replications} réplicas) ===")
    print(f"  Workers utilizados: {cfg.n_workers or multiprocessing.cpu_count()}")

    print("\n-- Volúmenes --")
    print(f"1ra: Agendados {fmt(agg['first_bookings'],0)} | Atendidos {fmt(agg['first_attended'],0)}")
    print(f"CNU 1ra: {fmt(agg['first_cnu_count'],0)}")
    print(f"Post ctrl bloqueado: {fmt(agg['post_ctrl_blocked'],0)}")
    print(f"Post: Ingresan {fmt(agg['post_entered'],0)} | Completan {fmt(agg['post_completed'],0)}")
    print(f"Post ctrl: Agendados {fmt(agg['post_ctrl_bookings'],0)} | Atendidos {fmt(agg['post_ctrl_attended'],0)}")
    print(f"Quirúrgico alterado: total={fmt(agg['quir_altered_total'],0)} "
          f"tto={fmt(agg['quir_altered_treatment'],0)} "
          f"anesth={fmt(agg['quir_altered_anesth'],0)}")

    print("\n-- Tiempos (días) --")
    print(f"TTS 1ra cerrada  : {fmt(agg['tts_first_closed_days_mean'])}")
    print(f"TTS 1ra atendida : {fmt(agg['tts_first_attended_days_mean'])}")
    print(f"  backlog  : {fmt(agg['tts_first_backlog_days_mean'])}")
    print(f"  proceso  : {fmt(agg['tts_first_process_days_mean'])}")
    print(f"TTS post         : {fmt(agg['tts_post_days_mean'])}")
    print(f"TTS total        : {fmt(agg['tts_full_days_mean'])}")

    print("\n-- Utilización (%) --")
    print(f"Agente: 1ra {fmt(agg['agent_util_first_pct'])} | post {fmt(agg['agent_util_post_pct'])}")
    print(f"Matrona: {fmt(agg['matrona_util_post_pct'])}")
    print(f"Especialistas: 1ra {fmt(agg['spec_util_first_pct'])} | post {fmt(agg['spec_util_post_pct'])}")

    print("\n-- Listas de espera al final --")
    print(f"1ra consulta : {fmt(agg['wl_first_end'],0)}")
    print(f"Control post : {fmt(agg['wl_control_end'],0)}")
    print(f"Total sistema: {fmt(agg['wl_total_end'],0)}")


# ========================
# Plots (sin cambios)
# ========================
def plot_one_run(run_result, agg=None):
    def mu(key, default=0.0):
        if agg and key in agg:
            return float(agg[key]['mean'])
        return float(run_result.get(key, default))

    plt.rcParams['figure.autolayout'] = True
    plt.rcParams['figure.figsize']    = (12, 5)

    ts1 = run_result.get("_ts_first", {})
    tsp = run_result.get("_ts_post",  {})

    if ts1.get("t"):
        x1 = [t / (60 * 24.0) for t in ts1["t"]]
        plt.figure()
        plt.plot(x1, ts1.get("cum_wait", []), label="Acum. en espera")
        plt.plot(x1, ts1.get("cum_att",  []), label="Acum. atendidos")
        plt.plot(x1, ts1.get("diff",     []), label="Diferencia")
        plt.xlabel("Día"); plt.ylabel("Pacientes")
        plt.title("Primera consulta: acumulados")
        plt.grid(True); plt.legend()

    if tsp.get("t"):
        xp = [t / (60 * 24.0) for t in tsp["t"]]
        plt.figure()
        plt.plot(xp, tsp.get("entered",    []), label="Post: ingresan")
        plt.plot(xp, tsp.get("completed",  []), label="Post: completan")
        plt.plot(xp, tsp.get("in_process", []), label="Post: en proceso")
        plt.xlabel("Día"); plt.ylabel("Pacientes")
        plt.title("Post-consulta: flujo")
        plt.grid(True); plt.legend()

    tsw = run_result.get("_ts_wl_week", {})
    if tsw.get("week"):
        plt.figure()
        plt.plot(tsw["week"], tsw["wl_first"],   label="1ra consulta")
        plt.plot(tsw["week"], tsw["wl_control"],  label="Control")
        plt.plot(tsw["week"], tsw["wl_preq"],     label="Prequirúrgico")
        plt.plot(tsw["week"], tsw["wl_quir"],     label="Quirúrgico")
        plt.xlabel("Semana"); plt.ylabel("Pacientes en espera")
        plt.title("Listas de espera por semana")
        plt.grid(True); plt.legend()

    plt.figure()
    labels = ["Agente 1ra", "Agente post", "Matrona", "Esp. 1ra", "Esp. post"]
    vals   = [mu("agent_util_first_pct"), mu("agent_util_post_pct"),
              mu("matrona_util_post_pct"), mu("spec_util_first_pct"), mu("spec_util_post_pct")]
    plt.bar(labels, vals)
    plt.ylabel("%"); plt.title("Utilización de recursos")
    plt.grid(True, axis="y"); plt.xticks(rotation=20, ha="right")

    plt.figure()
    labels = ["Espera 1ra", "Control", "Total"]
    vals   = [mu("wl_first_end"), mu("wl_control_end"), mu("wl_total_end")]
    plt.bar(labels, vals)
    plt.ylabel("Pacientes"); plt.title("Listas de espera — fin horizonte")
    plt.grid(True, axis="y")

    plt.show(block=True)


# ========================
# Main — versión paralelizada
# ========================
def main():
    cfg = CFG

    # Determinar número de workers
    n_workers = cfg.n_workers if cfg.n_workers > 0 else multiprocessing.cpu_count()
    n_workers = min(n_workers, cfg.replications)   # no más workers que réplicas

    log.info(
        "Iniciando simulación paralela: %d semanas, %d réplicas, %d workers",
        cfg.weeks_to_simulate, cfg.replications, n_workers,
    )

    # Serializar cfg una sola vez
    cfg_dict = _cfg_to_dict(cfg)

    # Construir lista de argumentos: (seed_offset, cfg_dict)
    tasks = [(r, cfg_dict) for r in range(cfg.replications)]

    # ── Ejecución paralela ──────────────────────────────────────────────────
    # ProcessPoolExecutor usa 'spawn' en Windows/macOS y 'fork' en Linux.
    # El guard if __name__ == "__main__" es OBLIGATORIO en Windows/macOS
    # para evitar que los workers relancen main() recursivamente.
    rep_results = []
    with concurrent.futures.ProcessPoolExecutor(max_workers=n_workers) as executor:
        futures = {executor.submit(_run_worker, task): task[0] for task in tasks}
        for future in concurrent.futures.as_completed(futures):
            seed = futures[future]
            try:
                result = future.result()
                rep_results.append(result)
                log.info("Réplica %d completada (%d/%d)", seed, len(rep_results), cfg.replications)
            except Exception as exc:
                log.error("Réplica %d falló: %s", seed, exc)

    if not rep_results:
        log.error("Ninguna réplica completó exitosamente.")
        return

    # Reordenar por seed para reproducibilidad del primer plot
    # (las futures completan en orden arbitrario)
    rep_results.sort(key=lambda r: r.get("validation_error_count", 0))

    agg = summarize(rep_results)

    for fn, path in [(export_summary_csv,       'kpi_summary.csv'),
                     (export_replications_csv,  'kpi_replications.csv'),
                     (export_summary_xlsx,       'kpi_summary.xlsx')]:
        try:
            fn(agg, path) if 'rep' not in path else fn(rep_results, path)
            log.info("Exportado: %s", path)
        except Exception as e:
            log.warning("No pudo exportar %s: %s", path, e)

    pretty_print(agg, cfg=cfg)
    plot_one_run(rep_results[0], agg=agg)


# ── Guard obligatorio para multiprocessing en Windows y macOS ───────────────
if __name__ == "__main__":
    main()
